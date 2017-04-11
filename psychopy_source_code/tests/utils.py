from os.path import abspath, basename, dirname, isfile, join as pjoin
import os.path
import shutil
import numpy as np
from psychopy import logging

try:
    from PIL import Image
except ImportError:
    import Image

try:
    import pytest
    usePytest=True
except Exception:
    usePytest=False

from pytest import skip

# define the path where to find testing data
# so tests could be ran from any location
TESTS_PATH = abspath(dirname(__file__))
TESTS_DATA_PATH = pjoin(TESTS_PATH, 'data')
TESTS_FONT = pjoin(TESTS_DATA_PATH, 'DejaVuSerif.ttf')

def compareScreenshot(fileName, win, crit=5.0):
    """Compare the current back buffer of the given window with the file

    Screenshots are stored and compared against the files under path
    kept in TESTS_DATA_PATH.  Thus specify relative path to that
    directory
    """
    #if we start this from a folder below run.py the data folder won't be found
    fileName = pjoin(TESTS_DATA_PATH, fileName)
    #get the frame from the window
    win.getMovieFrame(buffer='back')
    frame=win.movieFrames[-1]
    win.movieFrames=[]
    #if the file exists run a test, if not save the file
    if not isfile(fileName):
        frame.save(fileName, optimize=1)
        skip("Created %s" % basename(fileName))
    else:
        expected = Image.open(fileName)
        expDat = np.array(expected.getdata())
        imgDat = np.array(frame.getdata())
        rms = (((imgDat-expDat)**2).sum()/len(imgDat))**0.5
        filenameLocal = fileName.replace('.png','_local.png')
        if rms >= crit/2:
            #there was SOME discrepency
            logging.warning('PsychoPyTests: RMS=%.3g at threshold=%3.g'
                  % (rms, crit))
        if not rms<crit: #don't do `if rms>=crit because that doesn't catch rms=nan
            frame.save(filenameLocal, optimize=1)
            logging.warning('PsychoPyTests: Saving local copy into %s' % filenameLocal)
        assert rms<crit, \
            "RMS=%.3g at threshold=%.3g. Local copy in %s" % (rms, crit, filenameLocal)


def compareTextFiles(pathToActual, pathToCorrect, delim=None):
    """Compare the text of two files, ignoring EOL differences, and save a copy if they differ
    """
    if not os.path.isfile(pathToCorrect):
        logging.warning('There was no comparison ("correct") file available, saving current file as the comparison:%s' %pathToCorrect)
        foundComparisonFile=False
        shutil.copyfile(pathToActual,pathToCorrect)
        assert foundComparisonFile #deliberately raise an error to see the warning message
        return
    if delim is None:
        if pathToCorrect.endswith('.csv'):
            delim=','
        elif pathToCorrect.endswith(('.dlm', '.tsv')):
            delim='\t'

    try:
        #we have the necessary file
        txtActual = open(pathToActual, 'r').readlines()
        txtCorrect = open(pathToCorrect, 'r').readlines()
        assert len(txtActual)==len(txtCorrect), "The data file has the wrong number of lines"
        for lineN in range(len(txtActual)):
            if delim is None:
                #just compare the entire line
                assert lineActual==lineCorrect
            else:#word by word instead
                lineActual=txtActual[lineN].split(delim)
                lineCorrect=txtCorrect[lineN].split(delim)
                for wordN in range(len(lineActual)):
                    wordActual=lineActual[wordN]
                    wordCorrect=lineCorrect[wordN]
                    try:
                        wordActual=float(wordActual.lstrip('"[').strip(']"'))
                        wordCorrect=float(wordCorrect.lstrip('"[').strip(']"'))
                        # its not a whole well-formed list because .split(delim)
                        isFloat=True
                    except Exception:#stick with simple text if not a float value
                        isFloat=False
                        pass
                    if isFloat:
                        #to a default of 8 dp?
                        assert np.allclose(wordActual,wordCorrect), "Numeric values at (%i,%i) differ: %f != %f " \
                            %(lineN, wordN, wordActual, wordCorrect)
                    else:
                        if wordActual!=wordCorrect:
                            print('actual:')
                            print(repr(txtActual[lineN]))
                            print(lineActual)
                            print('expected:')
                            print(repr(txtCorrect[lineN]))
                            print(lineCorrect)
                        assert wordActual==wordCorrect, "Values at (%i,%i) differ: %s != %s " \
                            %(lineN, wordN, repr(wordActual), repr(wordCorrect))
    except AssertionError, err:
        pathToLocal, ext = os.path.splitext(pathToCorrect)
        pathToLocal = pathToLocal+'_local'+ext
        shutil.copyfile(pathToActual,pathToLocal)
        print("txtActual!=txtCorr: Saving local copy to %s" %pathToLocal)
        raise AssertionError, err

def compareXlsxFiles(pathToActual, pathToCorrect):
    from openpyxl.reader.excel import load_workbook
    # Make sure the file is there
    expBook = load_workbook(pathToCorrect)
    actBook = load_workbook(pathToActual)
    error=None

    for wsN, expWS in enumerate(expBook.worksheets):
        actWS = actBook.worksheets[wsN]
        for key, expVal in expWS._cells.items():
            actVal = actWS._cells[key].value
            expVal = expVal.value
            # intercept lists-of-floats, which might mismatch by rounding error
            isListableFloatable = False
            if str(expVal).startswith('['):
                expValList = eval(str(expVal))
                try:
                    expVal = np.array(expValList, dtype=float)
                    actVal = np.array(eval(str(actVal)), dtype=float) # should go through if expVal does...
                    isListableFloatable = True
                except Exception:
                    pass # non-list+float-able at this point = default
            #determine whether there will be errors
            try:
                # convert to float if possible and compare with a reasonable
                # (default) precision
                expVal = float(expVal)
                isFloatable=True
            except Exception:
                isFloatable=False
            if isListableFloatable:
                if not np.allclose(expVal, actVal):
                    error = "l+f Cell %s: %f != %f" %(key, expVal, actVal)
                    break
            elif isFloatable and abs(expVal-float(actVal))>0.0001:
                error = "f Cell %s: %f != %f" %(key, expVal, actVal)
                break
            elif not isFloatable and expVal!=actVal:
                error = "nf Cell %s: %s != %s" %(key, expVal, actVal)
                break
    if error:
        pathToLocal, ext = os.path.splitext(pathToCorrect)
        pathToLocal = pathToLocal+'_local'+ext
        shutil.copyfile(pathToActual,pathToLocal)
        logging.warning("xlsxActual!=xlsxCorr: Saving local copy to %s" %pathToLocal)
        raise IOError, error

_travisTesting = bool(str(os.environ.get('TRAVIS')).lower() == 'true')  # in Travis-CI testing

# Alternative skip_under_travis implementation;
# Seems fine, but Jon / Jeremy can decide to use it or loose it.
#
# skip_under_travis = pytest.mark.skipif(_travisTesting == True,
#                                       reason="Cannot be tested under Travis-CI")

def skip_under_travis(fn=None):
    """Skip if a test is executed under Travis testing environment
    Could also be used as a decorator (if argument provided) or
    unparametrized in the code
    """
    # TODO: ad-hoc check ATM -- there might be better ways
    if _travisTesting:
        skip, msg = pytest.skip, "Cannot be tested under Travis-CI"
        if fn is not None:
            def _inner():
                skip(msg)
            _inner.__name__ = fn.__name__
            return _inner
        else:
            skip(msg)
    else:
        return fn
