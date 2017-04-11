# -*- coding: utf-8 -*-
"""Routines for handling data structures and analysis
"""
# Part of the PsychoPy library
# Copyright (C) 2015 Jonathan Peirce
# Distributed under the terms of the GNU General Public License (GPL).

from __future__ import absolute_import

import pandas
import cPickle
import string
import sys
import os
import time
import copy
import numpy
from scipy import optimize, special
import inspect  # so that Handlers can find the script that called them
import codecs
import weakref
import re
import warnings
import collections
from distutils.version import StrictVersion

try:
    # import openpyxl
    import openpyxl
    if StrictVersion(openpyxl.__version__) >= StrictVersion('2.4.0'):
        # openpyxl moved get_column_letter to utils.cell
        from openpyxl.utils.cell import get_column_letter
    else:
        from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl = True
except ImportError:
    haveOpenpyxl = False

try:
    import xlrd
    haveXlrd = True
except ImportError:
    haveXlrd = False

from psychopy import logging
from psychopy.tools.arraytools import extendArr, shuffleArray
from psychopy.tools.fileerrortools import handleFileCollision
from psychopy.tools.filetools import openOutputFile, genDelimiter
import psychopy
from psychopy.contrib.quest import QuestObject  # used for QuestHandler
from psychopy.contrib.psi import PsiObject  # used for PsiHandler

_experiments = weakref.WeakValueDictionary()
_nonalphanumeric_re = re.compile(r'\W')  # will match all bad var name chars


class ExperimentHandler(object):
    """A container class for keeping track of multiple loops/handlers

    Useful for generating a single data file from an experiment with many
    different loops (e.g. interleaved staircases or loops within loops

    :usage:

        exp = data.ExperimentHandler(name="Face Preference",version='0.1.0')

    """

    def __init__(self,
                 name='',
                 version='',
                 extraInfo=None,
                 runtimeInfo=None,
                 originPath=None,
                 savePickle=True,
                 saveWideText=True,
                 dataFileName='',
                 autoLog=True):
        """
        :parameters:

            name : a string or unicode
                As a useful identifier later

            version : usually a string (e.g. '1.1.0')
                To keep track of which version of the experiment was run

            extraInfo : a dictionary
                Containing useful information about this run
                (e.g. {'participant':'jwp','gender':'m','orientation':90} )

            runtimeInfo : :class:`psychopy.info.RunTimeInfo`
                Containining information about the system as detected at
                runtime

            originPath : string or unicode
                The path and filename of the originating script/experiment
                If not provided this will be determined as the path of the
                calling script.

            dataFileName : string
                This is defined in advance and the file will be saved at any
                point that the handler is removed or discarded (unless
                .abort() had been called in advance).
                The handler will attempt to populate the file even in the
                event of a (not too serious) crash!

            savePickle : True (default) or False

            saveWideText : True (default) or False

            autoLog : True (default) or False
        """
        self.loops = []
        self.loopsUnfinished = []
        self.name = name
        self.version = version
        self.runtimeInfo = runtimeInfo
        if extraInfo is None:
            self.extraInfo = {}
        else:
            self.extraInfo = extraInfo
        self.originPath = originPath
        self.savePickle = savePickle
        self.saveWideText = saveWideText
        self.dataFileName = dataFileName
        self.thisEntry = {}
        self.entries = []  # chronological list of entries
        self._paramNamesSoFar = []
        self.dataNames = []  # names of all the data (eg. resp.keys)
        self.autoLog = autoLog
        if dataFileName in ['', None]:
            logging.warning('ExperimentHandler created with no dataFileName'
                            ' parameter. No data will be saved in the event '
                            'of a crash')
        else:
            # fail now if we fail at all!
            checkValidFilePath(dataFileName, makeValid=True)

    def __del__(self):
        if self.dataFileName not in ['', None]:
            if self.autoLog:
                logging.debug(
                    'Saving data for %s ExperimentHandler' % self.name)
            if self.savePickle == True:
                self.saveAsPickle(self.dataFileName)
            if self.saveWideText == True:
                self.saveAsWideText(self.dataFileName + '.csv', delim=',')

    def addLoop(self, loopHandler):
        """Add a loop such as a :class:`~psychopy.data.TrialHandler`
        or :class:`~psychopy.data.StairHandler`
        Data from this loop will be included in the resulting data files.
        """
        self.loops.append(loopHandler)
        self.loopsUnfinished.append(loopHandler)
        # keep the loop updated that is now owned
        loopHandler.setExp(self)

    def loopEnded(self, loopHandler):
        """Informs the experiment handler that the loop is finished and not to
        include its values in further entries of the experiment.

        This method is called by the loop itself if it ends its iterations,
        so is not typically needed by the user.
        """
        if loopHandler in self.loopsUnfinished:
            self.loopsUnfinished.remove(loopHandler)

    def _getAllParamNames(self):
        """Returns the attribute names of loop parameters (trialN etc)
        that the current set of loops contain, ready to build a wide-format
        data file.
        """
        names = copy.deepcopy(self._paramNamesSoFar)
        # get names (or identifiers) for all contained loops
        for thisLoop in self.loops:
            theseNames, vals = self._getLoopInfo(thisLoop)
            for name in theseNames:
                if name not in names:
                    names.append(name)
        return names

    def _getExtraInfo(self):
        """Get the names and vals from the extraInfo dict (if it exists)
        """
        if type(self.extraInfo) != dict:
            names = []
            vals = []
        else:
            names = self.extraInfo.keys()
            vals = self.extraInfo.values()
        return names, vals

    def _getLoopInfo(self, loop):
        """Returns the attribute names and values for the current trial
        of a particular loop. Does not return data inputs from the subject,
        only info relating to the trial execution.
        """
        names = []
        vals = []
        name = loop.name
        # standard attributes
        for attr in ('thisRepN', 'thisTrialN', 'thisN', 'thisIndex',
                     'stepSizeCurrent'):
            if hasattr(loop, attr):
                attrName = name + '.' + attr.replace('Current', '')
                # append the attribute name and the current value
                names.append(attrName)
                vals.append(getattr(loop, attr))
        # method of constants
        if hasattr(loop, 'thisTrial'):
            trial = loop.thisTrial
            if hasattr(trial, 'items'):
                # is a TrialList object or a simple dict
                for attr, val in trial.items():
                    if attr not in self._paramNamesSoFar:
                        self._paramNamesSoFar.append(attr)
                    names.append(attr)
                    vals.append(val)
        # single StairHandler
        elif hasattr(loop, 'intensities'):
            names.append(name + '.intensity')
            if len(loop.intensities) > 0:
                vals.append(loop.intensities[-1])
            else:
                vals.append(None)

        return names, vals

    def addData(self, name, value):
        """Add the data with a given name to the current experiment.

        Typically the user does not need to use this function; if you added
        your data to the loop and had already added the loop to the
        experiment then the loop will automatically inform the experiment
        that it has received data.

        Multiple data name/value pairs can be added to any given entry of
        the data file and is considered part of the same entry until the
        nextEntry() call is made.

        e.g.::

            # add some data for this trial
            exp.addData('resp.rt', 0.8)
            exp.addData('resp.key', 'k')
            # end of trial - move to next line in data output
            exp.nextEntry()
        """
        if name not in self.dataNames:
            self.dataNames.append(name)
        # could just copy() every value, but not always needed, so check:
        try:
            hash(value)
        except TypeError:
            # unhashable type (list, dict, ...) == mutable, so need a copy()
            value = copy.deepcopy(value)
        self.thisEntry[name] = value

    def nextEntry(self):
        """Calling nextEntry indicates to the ExperimentHandler that the
        current trial has ended and so further addData() calls correspond
        to the next trial.
        """
        this = self.thisEntry
        # fetch data from each (potentially-nested) loop
        for thisLoop in self.loopsUnfinished:
            names, vals = self._getLoopInfo(thisLoop)
            for n, name in enumerate(names):
                this[name] = vals[n]
        # add the extraInfo dict to the data
        if type(self.extraInfo) == dict:
            this.update(self.extraInfo)
        self.entries.append(this)
        self.thisEntry = {}

    def saveAsWideText(self, fileName, delim=None,
                       matrixOnly=False,
                       appendFile=False,
                       encoding='utf-8',
                       fileCollisionMethod='rename'):
        """Saves a long, wide-format text file, with one line representing
        the attributes and data for a single trial. Suitable for analysis
        in R and SPSS.

        If `appendFile=True` then the data will be added to the bottom of
        an existing file. Otherwise, if the file exists already it will
        be overwritten

        If `matrixOnly=True` then the file will not contain a header row,
        which can be handy if you want to append data to an existing file
        of the same format.

        encoding:
            The encoding to use when saving a the file. Defaults to `utf-8`.

        fileCollisionMethod:
            Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        """
        # set default delimiter if none given
        if delim is None:
            delim = genDelimiter(fileName)

        # create the file or send to stdout
        f = openOutputFile(
            fileName, append=appendFile, delim=delim,
            fileCollisionMethod=fileCollisionMethod, encoding=encoding)

        names = self._getAllParamNames()
        names.extend(self.dataNames)
        # names from the extraInfo dictionary
        names.extend(self._getExtraInfo()[0])
        # write a header line
        if not matrixOnly:
            for heading in names:
                f.write(u'%s%s' % (heading, delim))
            f.write('\n')
        # write the data for each entry

        for entry in self.entries:
            for name in names:
                entry.keys()
                if name in entry.keys():
                    ename = unicode(entry[name])
                    if ',' in ename or '\n' in ename:
                        fmt = u'"%s"%s'
                    else:
                        fmt = u'%s%s'
                    f.write(fmt % (entry[name], delim))
                else:
                    f.write(delim)
            f.write('\n')
        if f != sys.stdout:
            f.close()
        logging.info('saved data to %r' % f.name)

    def saveAsPickle(self, fileName, fileCollisionMethod='rename'):
        """Basically just saves a copy of self (with data) to a pickle file.

        This can be reloaded if necessary and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`
        """
        # Store the current state of self.savePickle and self.saveWideText
        # for later use:
        # We are going to set both to False before saving,
        # so PsychoPy won't try to save again after loading the pickled
        # .psydat file from disk.
        #
        # After saving, the initial state of self.savePickle and
        # self.saveWideText is restored.
        #
        # See
        # https://groups.google.com/d/msg/psychopy-dev/Z4m_UX88q8U/UGuh1eeyjMEJ
        savePickle = self.savePickle
        saveWideText = self.saveWideText

        self.savePickle = False
        self.saveWideText = False

        # otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName += '.psydat'

        f = openOutputFile(fileName, append=False,
                           fileCollisionMethod=fileCollisionMethod)
        cPickle.dump(self, f)
        f.close()
        logging.info('saved data to %s' % f.name)
        self.savePickle = savePickle
        self.saveWideText = saveWideText

    def abort(self):
        """Inform the ExperimentHandler that the run was aborted.

        Experiment handler will attempt automatically to save data
        (even in the event of a crash if possible). So if you quit your
        script early you may want to tell the Handler not to save out
        the data files for this run. This is the method that allows you
        to do that.
        """
        self.savePickle = False
        self.saveWideText = False


class TrialType(dict):
    """This is just like a dict, except that you can access keys with obj.key
    """

    def __getattribute__(self, name):
        try:  # to get attr from dict in normal way (passing self)
            return dict.__getattribute__(self, name)
        except AttributeError:
            try:
                return self[name]
            except KeyError:
                msg = "TrialType has no attribute (or key) \'%s\'"
                raise AttributeError(msg % name)


class _BaseTrialHandler(object):

    def setExp(self, exp):
        """Sets the ExperimentHandler that this handler is attached to

        Do NOT attempt to set the experiment using::

            trials._exp = myExperiment

        because it needs to be performed using the `weakref` module.
        """
        # need to use a weakref to avoid creating a circular reference that
        # prevents effective object deletion
        expId = id(exp)
        _experiments[expId] = exp
        self._exp = expId
        # origin will have been stored by the exp so don't store again:
        self.origin = None

    def getExp(self):
        """Return the ExperimentHandler that this handler is attached to,
        if any. Returns None if not attached
        """
        if self._exp is None or self._exp not in _experiments:
            return None
        else:
            return _experiments[self._exp]

    def _terminate(self):
        """Remove references to ourself in experiments and terminate the loop
        """
        # remove ourself from the list of unfinished loops in the experiment
        exp = self.getExp()
        if exp != None:
            exp.loopEnded(self)
        # and halt the loop
        raise StopIteration

    def saveAsPickle(self, fileName, fileCollisionMethod='rename'):
        """Basically just saves a copy of the handler (with data) to a
        pickle file.

        This can be reloaded if necessary and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`
        """
        if self.thisTrialN < 1 and self.thisRepN < 1:
            # if both are < 1 we haven't started
            if self.autoLog:
                logging.info('.saveAsPickle() called but no trials completed.'
                             ' Nothing saved')
            return -1
        # otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName += '.psydat'

        f = openOutputFile(fileName, append=False,
                           fileCollisionMethod=fileCollisionMethod)
        cPickle.dump(self, f)
        f.close()
        logging.info('saved data to %s' % f.name)

    def saveAsText(self, fileName,
                   stimOut=None,
                   dataOut=('n', 'all_mean', 'all_std', 'all_raw'),
                   delim=None,
                   matrixOnly=False,
                   appendFile=True,
                   summarised=True,
                   fileCollisionMethod='rename',
                   encoding='utf-8'):
        """
        Write a text file with the data and various chosen stimulus attributes

        :Parameters:

        fileName:
            will have .tsv appended and can include path info.

        stimOut:
            the stimulus attributes to be output. To use this you need to
            use a list of dictionaries and give here the names of dictionary
            keys that you want as strings

        dataOut:
            a list of strings specifying the dataType and the analysis to
            be performed,in the form `dataType_analysis`. The data can be
            any of the types that you added using trialHandler.data.add()
            and the analysis can be either 'raw' or most things in the
            numpy library, including; 'mean','std','median','max','min'...
            The default values will output the raw, mean and std of all
            datatypes found

        delim:
            allows the user to use a delimiter other than tab
            ("," is popular with file extension ".csv")

        matrixOnly:
            outputs the data with no header row or extraInfo attached

        appendFile:
            will add this output to the end of the specified file if
            it already exists

        fileCollisionMethod:
            Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        encoding:
            The encoding to use when saving a the file. Defaults to `utf-8`.

        """
        if stimOut is None:
            stimOut = []

        if self.thisTrialN < 1 and self.thisRepN < 1:
            # if both are < 1 we haven't started
            if self.autoLog:
                logging.info('TrialHandler.saveAsText called but no trials'
                             ' completed. Nothing saved')
            return -1

        dataArray = self._createOutputArray(stimOut=stimOut,
                                            dataOut=dataOut,
                                            matrixOnly=matrixOnly)

        # set default delimiter if none given
        if delim is None:
            delim = genDelimiter(fileName)

        # create the file or send to stdout
        f = openOutputFile(
            fileName, append=appendFile, delim=delim,
            fileCollisionMethod=fileCollisionMethod, encoding=encoding)

        # loop through lines in the data matrix
        for line in dataArray:
            for cellN, entry in enumerate(line):
                # surround in quotes to prevent effect of delimiter
                if delim in unicode(entry):
                    f.write(u'"%s"' % unicode(entry))
                else:
                    f.write(unicode(entry))
                if cellN < (len(line) - 1):
                    f.write(delim)
            f.write("\n")  # add an EOL at end of each line
        if f != sys.stdout:
            f.close()
            if self.autoLog:
                logging.info('saved data to %s' % f.name)

    def printAsText(self, stimOut=None,
                    dataOut=('all_mean', 'all_std', 'all_raw'),
                    delim='\t',
                    matrixOnly=False):
        """Exactly like saveAsText() except that the output goes
        to the screen instead of a file
        """
        if stimOut is None:
            stimOut = []
        self.saveAsText('stdout', stimOut, dataOut, delim, matrixOnly)

    def saveAsExcel(self, fileName, sheetName='rawData',
                    stimOut=None,
                    dataOut=('n', 'all_mean', 'all_std', 'all_raw'),
                    matrixOnly=False,
                    appendFile=True,
                    fileCollisionMethod='rename'):
        """
        Save a summary data file in Excel OpenXML format workbook
        (:term:`xlsx`) for processing in most spreadsheet packages.
        This format is compatible with versions of Excel (2007 or greater)
        and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files (see
        :func:`TrialHandler.saveAsText()` )
        that data can be stored in multiple named sheets within the file.
        So you could have a single file named after your experiment and
        then have one worksheet for each participant. Or you could have
        one file for each participant and then multiple sheets for
        repeated sessions etc.

        The file extension `.xlsx` will be added if not given already.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include
                relative or absolute path

            sheetName: string
                the name of the worksheet within the file

            stimOut: list of strings
                the attributes of the trial characteristics to be output.
                To use this you need to have provided a list of dictionaries
                specifying to trialList parameter of the TrialHandler and
                give here the names of strings specifying entries in that
                dictionary

            dataOut: list of strings
                specifying the dataType and the analysis to
                be performed, in the form `dataType_analysis`. The data
                can be any of the types that you added using
                trialHandler.data.add() and the analysis can be either
                'raw' or most things in the numpy library, including
                'mean','std','median','max','min'. e.g. `rt_max` will give
                a column of max reaction times across the trials assuming
                that `rt` values have been stored. The default values will
                output the raw, mean and std of all datatypes found.

            appendFile: True or False
                If False any existing file with this name will be
                overwritten. If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will
                be added to make it unique.

            fileCollisionMethod: string
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`
                This is ignored if ``append`` is ``True``.

        """
        if stimOut is None:
            stimOut = []

        if self.thisTrialN < 1 and self.thisRepN < 1:
            # if both are < 1 we haven't started
            if self.autoLog:
                logging.info('TrialHandler.saveAsExcel called but no '
                             'trials completed. Nothing saved')
            return -1

        # NB this was based on the limited documentation (1 page wiki) for
        # openpyxl v1.0
        if not haveOpenpyxl:
            raise ImportError('openpyxl is required for saving files in'
                              ' Excel (xlsx) format, but was not found.')
            # return -1

        # create the data array to be sent to the Excel file
        dataArray = self._createOutputArray(stimOut=stimOut,
                                            dataOut=dataOut,
                                            matrixOnly=matrixOnly)

        # import necessary subpackages - they are small so won't matter to do
        # it here
        from openpyxl.workbook import Workbook
        from openpyxl.reader.excel import load_workbook

        if not fileName.endswith('.xlsx'):
            fileName += '.xlsx'
        # create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook = False
        else:
            if not appendFile:
                # the file exists but we're not appending, will be overwritten
                fileName = handleFileCollision(fileName,
                                               fileCollisionMethod)
            wb = Workbook()  # create new workbook
            wb.properties.creator = 'PsychoPy' + psychopy.__version__
            newWorkbook = True

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title = sheetName
        else:
            ws = wb.create_sheet()
            ws.title = sheetName

        # loop through lines in the data matrix
        for lineN, line in enumerate(dataArray):
            if line is None:
                continue
            for colN, entry in enumerate(line):
                if entry is None:
                    entry = ''
                try:
                    # if it can convert to a number (from numpy) then do it
                    val = float(entry)
                except Exception:
                    val = unicode(entry)
                _cell = _getExcelCellName(col=colN, row=lineN)
                ws.cell(_cell).value = val

        wb.save(filename=fileName)

    def getOriginPathAndFile(self, originPath=None):
        """Attempts to determine the path of the script that created this
        data file and returns both the path to that script and its contents.
        Useful to store the entire experiment with the data.

        If originPath is provided (e.g. from Builder) then this is used
        otherwise the calling script is the originPath (fine from a
        standard python script).
        """
        # self.originPath and self.origin (the contents of the origin file)
        if originPath == -1:
            return -1, None  # the user wants to avoid storing this
        elif originPath is None or not os.path.isfile(originPath):
            try:
                originPath = inspect.getouterframes(
                    inspect.currentframe())[2][1]
                if self.autoLog:
                    logging.debug("Using %s as origin file" % originPath)
            except Exception:
                if self.autoLog:
                    logging.debug("Failed to find origin file using "
                                  "inspect.getouterframes")
                return '', ''
        if os.path.isfile(originPath):  # do we NOW have a path?
            origin = codecs.open(originPath, "r", encoding="utf-8").read()
        else:
            origin = None
        return originPath, origin


class TrialHandler(_BaseTrialHandler):
    """Class to handle trial sequencing and data storage.

    Calls to .next() will fetch the next trial object given to this handler,
    according to the method specified (random, sequential, fullRandom).
    Calls will raise a StopIteration error if trials have finished.

    See demo_trialHandler.py

    The psydat file format is literally just a pickled copy of the
    TrialHandler object that saved it. You can open it with::

            from psychopy.tools.filetools import fromFile
            dat = fromFile(path)

    Then you'll find that `dat` has the following attributes that
    """

    def __init__(self,
                 trialList,
                 nReps,
                 method='random',
                 dataTypes=None,
                 extraInfo=None,
                 seed=None,
                 originPath=None,
                 name='',
                 autoLog=True):
        """

        :Parameters:

            trialList: a simple list (or flat array) of dictionaries
                specifying conditions. This can be imported from an
                excel/csv file using :func:`~psychopy.data.importConditions`

            nReps: number of repeats for all conditions

            method: *'random',* 'sequential', or 'fullRandom'
                'sequential' obviously presents the conditions in the order
                they appear in the list. 'random' will result in a shuffle
                of the conditions on each repeat, but all conditions
                occur once before the second repeat etc. 'fullRandom'
                fully randomises the trials across repeats as well,
                which means you could potentially run all trials of
                one condition before any trial of another.

            dataTypes: (optional) list of names for data storage.
                e.g. ['corr','rt','resp']. If not provided then these
                will be created as needed during calls to
                :func:`~psychopy.data.TrialHandler.addData`

            extraInfo: A dictionary
                This will be stored alongside the data and usually
                describes the experiment and subject ID, date etc.

            seed: an integer
                If provided then this fixes the random number generator to
                use the same pattern of trials, by seeding its startpoint

            originPath: a string describing the location of the
                script / experiment file path. The psydat file format will
                store a copy of the experiment if possible. If
                `originPath==None` is provided here then the TrialHandler
                will still store a copy of the script where it was
                created. If `OriginPath==-1` then nothing will be stored.

        :Attributes (after creation):

            .data - a dictionary of numpy arrays, one for each data type
                stored

            .trialList - the original list of dicts, specifying the conditions

            .thisIndex - the index of the current trial in the original
                conditions list

            .nTotal - the total number of trials that will be run

            .nRemaining - the total number of trials remaining

            .thisN - total trials completed so far

            .thisRepN - which repeat you are currently on

            .thisTrialN - which trial number *within* that repeat

            .thisTrial - a dictionary giving the parameters of the current
                trial

            .finished - True/False for have we finished yet

            .extraInfo - the dictionary of extra info as given at beginning

            .origin - the contents of the script or builder experiment that
                created the handler

        """
        self.name = name
        self.autoLog = autoLog

        if trialList in (None, []):  # user wants an empty trialList
            # which corresponds to a list with a single empty entry
            self.trialList = [None]
        # user has hopefully specified a filename
        elif isinstance(trialList, basestring) and os.path.isfile(trialList):
            # import conditions from that file
            self.trialList = importConditions(trialList)
        else:
            self.trialList = trialList
        # convert any entry in the TrialList into a TrialType object (with
        # obj.key or obj[key] access)
        for n, entry in enumerate(self.trialList):
            if type(entry) == dict:
                self.trialList[n] = TrialType(entry)
        self.nReps = int(nReps)
        self.nTotal = self.nReps * len(self.trialList)
        self.nRemaining = self.nTotal  # subtract 1 each trial
        self.method = method
        self.thisRepN = 0  # records which repetition or pass we are on
        self.thisTrialN = -1  # records trial number within this repetition
        self.thisN = -1
        self.thisIndex = 0  # index of current trial in the conditions list
        self.thisTrial = []
        self.finished = False
        self.extraInfo = extraInfo
        self.seed = seed
        # create dataHandler
        self.data = DataHandler(trials=self)
        if dataTypes != None:
            self.data.addDataType(dataTypes)
        self.data.addDataType('ran')
        self.data['ran'].mask = False  # this is a bool; all entries are valid
        self.data.addDataType('order')
        # generate stimulus sequence
        if self.method in ['random', 'sequential', 'fullRandom']:
            self.sequenceIndices = self._createSequence()
        else:
            self.sequenceIndices = []

        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None  # the experiment handler that owns me!

    def __iter__(self):
        return self

    def __repr__(self):
        """prints a more verbose version of self as string
        """
        return self.__str__(verbose=True)

    def __str__(self, verbose=False):
        """string representation of the object
        """
        strRepres = 'psychopy.data.{}(\n'.format(self.__class__.__name__)
        attribs = dir(self)

        # data first, then all others
        try:
            data = self.data
        except Exception:
            data = None
        if data:
            strRepres += str('\tdata=')
            strRepres += str(data) + '\n'

        for thisAttrib in attribs:
            # can handle each attribute differently
            if 'instancemethod' in str(type(getattr(self, thisAttrib))):
                # this is a method
                continue
            elif thisAttrib[0] == '_':
                # the attrib is private
                continue
            elif thisAttrib == 'data':
                # we handled this first
                continue
            elif len(str(getattr(self, thisAttrib))) > 20 and not verbose:
                # just give type of LONG public attribute
                strRepres += str('\t' + thisAttrib + '=')
                strRepres += str(type(getattr(self, thisAttrib))) + '\n'
            else:
                # give the complete contents of attribute
                strRepres += str('\t' + thisAttrib + '=')
                strRepres += str(getattr(self, thisAttrib)) + '\n'

        strRepres += ')'
        return strRepres

    def _createSequence(self):
        """Pre-generates the sequence of trial presentations
        (for non-adaptive methods). This is called automatically when
        the TrialHandler is initialised so doesn't need an explicit call
        from the user.

        The returned sequence has form indices[stimN][repN]
        Example: sequential with 6 trialtypes (rows), 5 reps (cols), returns:
            [[0 0 0 0 0]
             [1 1 1 1 1]
             [2 2 2 2 2]
             [3 3 3 3 3]
             [4 4 4 4 4]
             [5 5 5 5 5]]
        These 30 trials will be returned by .next() in the order:
            0, 1, 2, 3, 4, 5,   0, 1, 2, ...  ... 3, 4, 5

        To add a new type of sequence (as of v1.65.02):
        - add the sequence generation code here
        - adjust "if self.method in [ ...]:" in both __init__ and .next()
        - adjust allowedVals in experiment.py -> shows up in DlgLoopProperties
        Note that users can make any sequence whatsoever outside of PsychoPy,
        and specify sequential order; any order is possible this way.
        """
        # create indices for a single rep
        indices = numpy.asarray(self._makeIndices(self.trialList), dtype=int)

        if self.method == 'random':
            sequenceIndices = []
            seed = self.seed
            for thisRep in range(self.nReps):
                thisRepSeq = shuffleArray(indices.flat, seed=seed).tolist()
                seed = None  # so that we only seed the first pass through!
                sequenceIndices.append(thisRepSeq)
            sequenceIndices = numpy.transpose(sequenceIndices)
        elif self.method == 'sequential':
            sequenceIndices = numpy.repeat(indices, self.nReps, 1)
        elif self.method == 'fullRandom':
            # indices*nReps, flatten, shuffle, unflatten; only use seed once
            sequential = numpy.repeat(indices, self.nReps, 1)  # = sequential
            randomFlat = shuffleArray(sequential.flat, seed=self.seed)
            sequenceIndices = numpy.reshape(
                randomFlat, (len(indices), self.nReps))
        if self.autoLog:
            msg = 'Created sequence: %s, trialTypes=%d, nReps=%i, seed=%s'
            vals = (self.method, len(indices), self.nReps, str(self.seed))
            logging.exp(msg % vals)
        return sequenceIndices

    def _makeIndices(self, inputArray):
        """
        Creates an array of tuples the same shape as the input array
        where each tuple contains the indices to itself in the array.

        Useful for shuffling and then using as a reference.
        """
        # make sure its an array of objects (can be strings etc)
        inputArray = numpy.asarray(inputArray, 'O')
        # get some simple variables for later
        dims = inputArray.shape
        dimsProd = numpy.product(dims)
        dimsN = len(dims)
        dimsList = range(dimsN)
        listOfLists = []
        # this creates space for an array of any objects
        arrayOfTuples = numpy.ones(dimsProd, 'O')

        # for each dimension create list of its indices (using modulo)
        for thisDim in dimsList:
            prevDimsProd = numpy.product(dims[:thisDim])
            # NB this means modulus in python
            thisDimVals = numpy.arange(dimsProd) / prevDimsProd % dims[thisDim]
            listOfLists.append(thisDimVals)

        # convert to array
        indexArr = numpy.asarray(listOfLists)
        for n in range(dimsProd):
            arrayOfTuples[n] = tuple((indexArr[:, n]))
        return (numpy.reshape(arrayOfTuples, dims)).tolist()

    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; thisTrial, thisTrialN and thisIndex
        If the trials have ended this method will raise a StopIteration error.
        This can be handled with code such as::

            trials = data.TrialHandler(.......)
            for eachTrial in trials:  # automatically stops when done
                # do stuff

        or::

            trials = data.TrialHandler(.......)
            while True:  # ie forever
                try:
                    thisTrial = trials.next()
                except StopIteration:  # we got a StopIteration error
                    break #break out of the forever loop
                # do stuff here for the trial
        """
        # update pointer for next trials
        self.thisTrialN += 1  # number of trial this pass
        self.thisN += 1  # number of trial in total
        self.nRemaining -= 1
        if self.thisTrialN == len(self.trialList):
            # start a new repetition
            self.thisTrialN = 0
            self.thisRepN += 1
        if self.thisRepN >= self.nReps:
            # all reps complete
            self.thisTrial = []
            self.finished = True

        if self.finished == True:
            self._terminate()

        # fetch the trial info
        if self.method in ('random', 'sequential', 'fullRandom'):
            self.thisIndex = self.sequenceIndices[
                self.thisTrialN][self.thisRepN]
            self.thisTrial = self.trialList[self.thisIndex]
            self.data.add('ran', 1)
            self.data.add('order', self.thisN)
        if self.autoLog:
            msg = 'New trial (rep=%i, index=%i): %s'
            vals = (self.thisRepN, self.thisTrialN, self.thisTrial)
            logging.exp(msg % vals, obj=self.thisTrial)
        return self.thisTrial

    def getFutureTrial(self, n=1):
        """Returns the condition for n trials into the future,
        without advancing the trials. A negative n returns a previous (past)
        trial. Returns 'None' if attempting to go beyond the last trial.
        """
        # check that we don't go out of bounds for either positive or negative
        if n > self.nRemaining or self.thisN + n < 0:
            return None
        seqs = numpy.array(self.sequenceIndices).transpose().flat
        condIndex = seqs[self.thisN + n]
        return self.trialList[condIndex]

    def getEarlierTrial(self, n=-1):
        """Returns the condition information from n trials previously.
        Useful for comparisons in n-back tasks. Returns 'None' if trying
        to access a trial prior to the first.
        """
        # treat positive offset values as equivalent to negative ones:
        return self.getFutureTrial(-abs(n))

    def _createOutputArray(self, stimOut, dataOut, delim=None,
                           matrixOnly=False):
        """Does the leg-work for saveAsText and saveAsExcel.
        Combines stimOut with ._parseDataOutput()
        """
        if (stimOut == [] and
                len(self.trialList) and
                hasattr(self.trialList[0], 'keys')):
            stimOut = self.trialList[0].keys()
            # these get added somewhere (by DataHandler?)
            if 'n' in stimOut:
                stimOut.remove('n')
            if 'float' in stimOut:
                stimOut.remove('float')

        lines = []
        # parse the dataout section of the output
        dataOut, dataAnal, dataHead = self._createOutputArrayData(dataOut)
        if not matrixOnly:
            thisLine = []
            lines.append(thisLine)
            # write a header line
            for heading in stimOut + dataHead:
                if heading == 'ran_sum':
                    heading = 'n'
                elif heading == 'order_raw':
                    heading = 'order'
                thisLine.append(heading)

        # loop through stimuli, writing data
        for stimN in range(len(self.trialList)):
            thisLine = []
            lines.append(thisLine)
            # first the params for this stim (from self.trialList)
            for heading in stimOut:
                thisLine.append(self.trialList[stimN][heading])

            # then the data for this stim (from self.data)
            for thisDataOut in dataOut:
                # make a string version of the data and then format it
                tmpData = dataAnal[thisDataOut][stimN]
                if hasattr(tmpData, 'tolist'):  # is a numpy array
                    strVersion = unicode(tmpData.tolist())
                    # for numeric data replace None with a blank cell
                    if tmpData.dtype.kind not in ['SaUV']:
                        strVersion = strVersion.replace('None', '')
                elif tmpData in [None, 'None']:
                    strVersion = ''
                else:
                    strVersion = unicode(tmpData)

                if strVersion == '()':
                    # 'no data' in masked array should show as "--"
                    strVersion = "--"
                # handle list of values (e.g. rt_raw )
                if (len(strVersion) and
                        strVersion[0] in '[(' and
                        strVersion[-1] in '])'):
                    strVersion = strVersion[1:-1]  # skip first and last chars
                # handle lists of lists (e.g. raw of multiple key presses)
                if (len(strVersion) and
                        strVersion[0] in '[(' and
                        strVersion[-1] in '])'):
                    tup = eval(strVersion)  # convert back to a tuple
                    for entry in tup:
                        # contents of each entry is a list or tuple so keep in
                        # quotes to avoid probs with delim
                        thisLine.append(unicode(entry))
                else:
                    thisLine.extend(strVersion.split(','))

        # add self.extraInfo
        if (self.extraInfo != None) and not matrixOnly:
            lines.append([])
            # give a single line of space and then a heading
            lines.append(['extraInfo'])
            for key, value in self.extraInfo.items():
                lines.append([key, value])
        return lines

    def _createOutputArrayData(self, dataOut):
        """This just creates the dataOut part of the output matrix.
        It is called by _createOutputArray() which creates the header
        line and adds the stimOut columns
        """
        dataHead = []  # will store list of data headers
        dataAnal = dict([])  # will store data that has been analyzed
        if type(dataOut) == str:
            # don't do list convert or we get a list of letters
            dataOut = [dataOut]
        elif type(dataOut) != list:
            dataOut = list(dataOut)

        # expand any 'all' dataTypes to be full list of available dataTypes
        allDataTypes = self.data.keys()
        # treat these separately later
        allDataTypes.remove('ran')
        # ready to go through standard data types
        dataOutNew = []
        for thisDataOut in dataOut:
            if thisDataOut == 'n':
                # n is really just the sum of the ran trials
                dataOutNew.append('ran_sum')
                continue  # no need to do more with this one
            # then break into dataType and analysis
            dataType, analType = string.rsplit(thisDataOut, '_', 1)
            if dataType == 'all':
                dataOutNew.extend(
                    [key + "_" + analType for key in allDataTypes])
                if 'order_mean' in dataOutNew:
                    dataOutNew.remove('order_mean')
                if 'order_std' in dataOutNew:
                    dataOutNew.remove('order_std')
            else:
                dataOutNew.append(thisDataOut)
        dataOut = dataOutNew
        # sort so all datatypes come together, rather than all analtypes
        dataOut.sort()

        # do the various analyses, keeping track of fails (e.g. mean of a
        # string)
        dataOutInvalid = []
        # add back special data types (n and order)
        if 'ran_sum' in dataOut:
            # move n to the first column
            dataOut.remove('ran_sum')
            dataOut.insert(0, 'ran_sum')
        if 'order_raw' in dataOut:
            # move order_raw to the second column
            dataOut.remove('order_raw')
            dataOut.append('order_raw')
        # do the necessary analysis on the data
        for thisDataOutN, thisDataOut in enumerate(dataOut):
            dataType, analType = string.rsplit(thisDataOut, '_', 1)
            if not dataType in self.data:
                # that analysis can't be done
                dataOutInvalid.append(thisDataOut)
                continue
            thisData = self.data[dataType]

            # set the header
            dataHead.append(dataType + '_' + analType)
            # analyse thisData using numpy module
            if analType in dir(numpy):
                try:
                    # will fail if we try to take mean of a string for example
                    if analType == 'std':
                        thisAnal = numpy.std(thisData, axis=1, ddof=0)
                        # normalise by N-1 instead. This should work by
                        # setting ddof=1 but doesn't as of 08/2010 (because
                        # of using a masked array?)
                        N = thisData.shape[1]
                        if N == 1:
                            thisAnal *= 0  # prevent a divide-by-zero error
                        else:
                            sqrt = numpy.sqrt
                            thisAnal = thisAnal * sqrt(N) / sqrt(N - 1)
                    else:
                        thisAnal = eval("numpy.%s(thisData,1)" % analType)
                except Exception:
                    # that analysis doesn't work
                    dataHead.remove(dataType + '_' + analType)
                    dataOutInvalid.append(thisDataOut)
                    continue  # to next analysis
            elif analType == 'raw':
                thisAnal = thisData
            else:
                raise AttributeError('You can only use analyses from numpy')
            # add extra cols to header if necess
            if len(thisAnal.shape) > 1:
                for n in range(thisAnal.shape[1] - 1):
                    dataHead.append("")
            dataAnal[thisDataOut] = thisAnal

        # remove invalid analyses (e.g. average of a string)
        for invalidAnal in dataOutInvalid:
            dataOut.remove(invalidAnal)
        return dataOut, dataAnal, dataHead

    def saveAsWideText(self, fileName,
                       delim=None,
                       matrixOnly=False,
                       appendFile=True,
                       encoding='utf-8',
                       fileCollisionMethod='rename'):
        """Write a text file with the session, stimulus, and data values
        from each trial in chronological order. Also, return a
        pandas.DataFrame containing same information as the file.

        That is, unlike 'saveAsText' and 'saveAsExcel':
         - each row comprises information from only a single trial.
         - no summarizing is done (such as collapsing to produce mean and
           standard deviation values across trials).

        This 'wide' format, as expected by R for creating dataframes, and
        various other analysis programs, means that some information must
        be repeated on every row.

        In particular, if the trialHandler's 'extraInfo' exists, then each
        entry in there occurs in every row. In builder, this will include
        any entries in the 'Experiment info' field of the
        'Experiment settings' dialog. In Coder, this information can be
        set using something like::

            myTrialHandler.extraInfo = {'SubjID': 'Joan Smith',
                                        'Group': 'Control'}

        :Parameters:

            fileName:
                if extension is not specified, '.csv' will be appended
                if the delimiter is ',', else '.tsv' will be appended.
                Can include path info.

            delim:
                allows the user to use a delimiter other than the default
                tab ("," is popular with file extension ".csv")

            matrixOnly:
                outputs the data with no header row.

            appendFile:
                will add this output to the end of the specified file if
                it already exists.

            fileCollisionMethod:
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`

            encoding:
                The encoding to use when saving a the file.
                Defaults to `utf-8`.

        """
        if self.thisTrialN < 1 and self.thisRepN < 1:
            # if both are < 1 we haven't started
            logging.info('TrialHandler.saveAsWideText called but no '
                         'trials completed. Nothing saved')
            return -1

        # set default delimiter if none given
        if delim is None:
            delim = genDelimiter(fileName)

        # create the file or send to stdout
        f = openOutputFile(
            fileName, append=appendFile, delim=delim,
            fileCollisionMethod=fileCollisionMethod, encoding=encoding)

        # collect parameter names related to the stimuli:
        if self.trialList[0]:
            header = self.trialList[0].keys()
        else:
            header = []
        # and then add parameter names related to data (e.g. RT)
        header.extend(self.data.dataTypes)
        # get the extra 'wide' parameter names into the header line:
        header.insert(0, "TrialNumber")
        # this is wide format, so we want fixed information
        # (e.g. subject ID, date, etc) repeated every line if it exists:
        if self.extraInfo is not None:
            for key in self.extraInfo:
                header.insert(0, key)
        df = pandas.DataFrame(columns=header)

        # loop through each trial, gathering the actual values:
        dataOut = []
        trialCount = 0
        # total number of trials = number of trialtypes * number of
        # repetitions:

        repsPerType = {}
        for rep in range(self.nReps):
            for trialN in range(len(self.trialList)):
                # find out what trial type was on this trial
                trialTypeIndex = self.sequenceIndices[trialN, rep]
                # determine which repeat it is for this trial
                if trialTypeIndex not in repsPerType.keys():
                    repsPerType[trialTypeIndex] = 0
                else:
                    repsPerType[trialTypeIndex] += 1
                # what repeat are we on for this trial type?
                trep = repsPerType[trialTypeIndex]

                # create a dictionary representing each trial:
                nextEntry = {}

                # add a trial number so the original order of the data can
                # always be recovered if sorted during analysis:
                trialCount += 1

                # now collect the value from each trial of vars in header:
                for prmName in header:
                    # the header includes both trial and data variables, so
                    # need to check before accessing:
                    tti = trialTypeIndex
                    if self.trialList[tti] and prmName in self.trialList[tti]:
                        nextEntry[prmName] = self.trialList[tti][prmName]
                    elif prmName in self.data:
                        nextEntry[prmName] = self.data[prmName][tti][trep]
                    elif self.extraInfo != None and prmName in self.extraInfo:
                        nextEntry[prmName] = self.extraInfo[prmName]
                    else:
                        # allow a null value if this parameter wasn't
                        # explicitly stored on this trial:
                        if prmName == "TrialNumber":
                            nextEntry[prmName] = trialCount
                        else:
                            nextEntry[prmName] = ''

                # store this trial's data
                dataOut.append(nextEntry)
                df = df.append(nextEntry, ignore_index=True)

        if not matrixOnly:
            # write the header row:
            nextLine = ''
            for prmName in header:
                nextLine = nextLine + prmName + delim
            # remove the final orphaned tab character
            f.write(nextLine[:-1] + '\n')

        # write the data matrix:
        for trial in dataOut:
            nextLine = ''
            for prmName in header:
                nextLine = nextLine + unicode(trial[prmName]) + delim
            # remove the final orphaned tab character
            nextLine = nextLine[:-1]
            f.write(nextLine + '\n')

        if f != sys.stdout:
            f.close()
            logging.info('saved wide-format data to %s' % f.name)

        # Converts numbers to numeric, such as float64, boolean to bool.
        # Otherwise they all are "object" type, i.e. strings
        df = df.convert_objects()
        return df

    def addData(self, thisType, value, position=None):
        """Add data for the current trial
        """
        self.data.add(thisType, value, position=None)
        if self.getExp() != None:  # update the experiment handler too
            self.getExp().addData(thisType, value)


class TrialHandler2(_BaseTrialHandler):
    """Class to handle trial sequencing and data storage.

    Calls to .next() will fetch the next trial object given to this handler,
    according to the method specified (random, sequential, fullRandom).
    Calls will raise a StopIteration error if trials have finished.

    See demo_trialHandler.py

    The psydat file format is literally just a pickled copy of the
    TrialHandler object that saved it. You can open it with::

            from psychopy.tools.filetools import fromFile
            dat = fromFile(path)

    Then you'll find that `dat` has the following attributes that
    """

    def __init__(self,
                 trialList,
                 nReps,
                 method='random',
                 dataTypes=None,
                 extraInfo=None,
                 seed=None,
                 originPath=None,
                 name='',
                 autoLog=True):
        """

        :Parameters:

            trialList: a simple list (or flat array) of dictionaries
                specifying conditions. This can be imported from an
                excel / csv file using :func:`~psychopy.data.importConditions`

            nReps: number of repeats for all conditions

            method: *'random',* 'sequential', or 'fullRandom'
                'sequential' obviously presents the conditions in the order
                they appear in the list. 'random' will result in a shuffle
                of the conditions on each repeat, but all conditions occur
                once before the second repeat etc. 'fullRandom' fully
                randomises the trials across repeats as well, which means
                you could potentially run all trials of one condition
                before any trial of another.

            dataTypes: (optional) list of names for data storage.
                e.g. ['corr','rt','resp']. If not provided then these
                will be created as needed during calls to
                :func:`~psychopy.data.TrialHandler.addData`

            extraInfo: A dictionary
                This will be stored alongside the data and usually describes
                the experiment and subject ID, date etc.

            seed: an integer
                If provided then this fixes the random number generator to
                use the same pattern of trials, by seeding its startpoint.

            originPath: a string describing the location of the script /
                experiment file path. The psydat file format will store a
                copy of the experiment if possible. If `originPath==None`
                is provided here then the TrialHandler will still store a
                copy of the script where it was
                created. If `OriginPath==-1` then nothing will be stored.

        :Attributes (after creation):

            .data - a dictionary of numpy arrays, one for each data type
                stored

            .trialList - the original list of dicts, specifying the conditions

            .thisIndex - the index of the current trial in the original
                conditions list

            .nTotal - the total number of trials that will be run

            .nRemaining - the total number of trials remaining

            .thisN - total trials completed so far

            .thisRepN - which repeat you are currently on

            .thisTrialN - which trial number *within* that repeat

            .thisTrial - a dictionary giving the parameters of the current
                trial

            .finished - True/False for have we finished yet

            .extraInfo - the dictionary of extra info as given at beginning

            .origin - the contents of the script or builder experiment that
                created the handler

        """
        self.name = name
        self.autoLog = autoLog

        if trialList in [None, [None], []]:  # user wants an empty trialList
            # which corresponds to a list with a single empty entry
            self.trialList = [None]
            self.columns = []
        # user has hopefully specified a filename
        elif isinstance(trialList, basestring) and os.path.isfile(trialList):
            # import conditions from that file
            self.trialList, self.columns = importConditions(trialList, True)
        else:
            self.trialList = trialList
            self.columns = trialList[0].keys()
        # convert any entry in the TrialList into a TrialType object (with
        # obj.key or obj[key] access)
        for n, entry in enumerate(self.trialList):
            if type(entry) == dict:
                self.trialList[n] = TrialType(entry)
        self.nReps = int(nReps)
        self.nTotal = self.nReps * len(self.trialList)
        self.nRemaining = self.nTotal  # subtract 1 each trial
        self.remainingIndices = []
        self.prevIndices = []
        self.method = method
        self.thisRepN = 0  # records which repetition or pass we are on
        self.thisTrialN = -1  # records trial number within this repetition
        self.thisN = -1
        self.thisIndex = None  # index of current trial in the conditions list
        self.thisTrial = {}
        self.finished = False
        self.extraInfo = extraInfo
        self.seed = seed
        self._rng = numpy.random.RandomState(seed=seed)

        # store a list of dicts, convert to pandas.DataFrame on access
        self._data = []

        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None  # the experiment handler that owns me!

    def __iter__(self):
        return self

    def __repr__(self):
        """prints a more verbose version of self as string
        """
        return self.__str__(verbose=True)

    def __str__(self, verbose=False):
        """string representation of the object
        """
        strRepres = 'psychopy.data.{}(\n'.format(self.__class__.__name__)
        attribs = dir(self)
        # data first, then all others
        try:
            data = self.data
        except Exception:
            strRepres += '\t(no data)\n'
        else:
            strRepres += str('\tdata=')
            strRepres += str(data) + '\n'
        for thisAttrib in attribs:
            # can handle each attribute differently
            if 'instancemethod' in str(type(getattr(self, thisAttrib))):
                # this is a method
                continue
            elif thisAttrib[0] == '_':
                # the attrib is private
                continue
            elif thisAttrib == 'data':
                # we handled this first
                continue
            elif (len(str(getattr(self, thisAttrib))) > 20 and
                    not verbose):
                # just give type of LONG public attribute
                strRepres += str('\t' + thisAttrib + '=')
                strRepres += str(type(getattr(self, thisAttrib))) + '\n'
            else:
                # give the complete contents of attribute
                strRepres += str('\t' + thisAttrib + '=')
                strRepres += str(getattr(self, thisAttrib)) + '\n'
        strRepres += ')'
        return strRepres

    @property
    def data(self):
        """Returns a pandas.DataFrame of the trial data so far
        Read only attribute - you can't directly modify TrialHandler.data

        Note that data are stored internally as a list of dictionaries,
        one per trial. These are converted to a DataFrame on access.
        """
        return pandas.DataFrame(self._data)

    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; thisTrial, thisTrialN and thisIndex
        If the trials have ended this method will raise a StopIteration error.
        This can be handled with code such as::

            trials = data.TrialHandler(.......)
            for eachTrial in trials:  # automatically stops when done
                # do stuff

        or::

            trials = data.TrialHandler(.......)
            while True:  # ie forever
                try:
                    thisTrial = trials.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial
        """
        # update pointer for next trials
        self.thisTrialN += 1  # number of trial this pass
        self.thisN += 1  # number of trial in total
        self.nRemaining -= 1
        if self.thisIndex is not None:
            self.prevIndices.append(self.thisIndex)

        # thisRepN has exceeded nReps
        if self.remainingIndices == []:
            # we've just started, or just starting a new repeat
            sequence = range(len(self.trialList))
            if (self.method == 'fullRandom' and
                    self.thisN < (self.nReps * len(self.trialList))):
                # we've only just started on a fullRandom sequence
                sequence *= self.nReps
                self._rng.shuffle(sequence)
                self.remainingIndices = sequence
            elif (self.method in ('sequential', 'random') and
                    self.thisRepN < self.nReps):
                # start a new repetition
                self.thisTrialN = 0
                self.thisRepN += 1
                if self.method == 'random':
                    self._rng.shuffle(sequence)  # shuffle in-place
                self.remainingIndices = sequence
            else:
                # we've finished
                self.finished = True
                self._terminate()  # raises Stop (code won't go beyond here)

        # fetch the trial info
        if len(self.trialList) == 0:
            self.thisIndex = 0
            self.thisTrial = {}
        else:
            self.thisIndex = self.remainingIndices.pop(0)
            # if None then use empty dict
            thisTrial = self.trialList[self.thisIndex] or {}
            self.thisTrial = copy.copy(thisTrial)
        # for fullRandom check how many times this has come up before
        if self.method == 'fullRandom':
            self.thisRepN = self.prevIndices.count(self.thisIndex)

        # update data structure with new info
        self._data.append(self.thisTrial)  # update the data list of dicts
        self.addData('thisN', self.thisN)
        self.addData('thisTrialN', self.thisTrialN)
        self.addData('thisRepN', self.thisRepN)
        if self.autoLog:
            msg = 'New trial (rep=%i, index=%i): %s'
            vals = (self.thisRepN, self.thisTrialN, self.thisTrial)
            logging.exp(msg % vals, obj=self.thisTrial)
        return self.thisTrial

    def getFutureTrial(self, n=1):
        """Returns the condition for n trials into the future, without
        advancing the trials. Returns 'None' if attempting to go beyond
        the last trial.
        """
        # check that we don't go out of bounds for either positive or negative
        # offsets:
        if n > self.nRemaining or self.thisN + n < 0:
            return None
        seqs = numpy.array(self.sequenceIndices).transpose().flat
        condIndex = seqs[self.thisN + n]
        return self.trialList[condIndex]

    def getEarlierTrial(self, n=-1):
        """Returns the condition information from n trials previously.
        Useful for comparisons in n-back tasks. Returns 'None' if trying
        to access a trial prior to the first.
        """
        # treat positive offset values as equivalent to negative ones:
        if n > 0:
            n = n * -1
        return self.getFutureTrial(n)

    def saveAsWideText(self, fileName,
                       delim=None,
                       matrixOnly=False,
                       appendFile=True,
                       encoding='utf-8',
                       fileCollisionMethod='rename'):
        """Write a text file with the session, stimulus, and data values
        from each trial in chronological order. Also, return a
        pandas.DataFrame containing same information as the file.

        That is, unlike 'saveAsText' and 'saveAsExcel':
         - each row comprises information from only a single trial.
         - no summarising is done (such as collapsing to produce mean and
           standard deviation values across trials).

        This 'wide' format, as expected by R for creating dataframes, and
        various other analysis programs, means that some information must
        be repeated on every row.

        In particular, if the trialHandler's 'extraInfo' exists, then each
        entry in there occurs in every row. In builder, this will include
        any entries in the 'Experiment info' field of the
        'Experiment settings' dialog. In Coder, this information can be set
        using something like::

            myTrialHandler.extraInfo = {'SubjID': 'Joan Smith',
                                        'Group': 'Control'}

        :Parameters:

            fileName:
                if extension is not specified, '.csv' will be appended if
                the delimiter is ',', else '.tsv' will be appended.
                Can include path info.

            delim:
                allows the user to use a delimiter other than the default
                tab ("," is popular with file extension ".csv")

            matrixOnly:
                outputs the data with no header row.

            appendFile:
                will add this output to the end of the specified file if
                it already exists.

            fileCollisionMethod:
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`

            encoding:
                The encoding to use when saving a the file.
                Defaults to `utf-8`.

        """
        if self.thisTrialN < 1 and self.thisRepN < 1:
            # if both are < 1 we haven't started
            logging.info('TrialHandler.saveAsWideText called but no '
                         'trials completed. Nothing saved')
            return -1

        # set default delimiter if none given
        if delim is None:
            delim = genDelimiter(fileName)

        # create the file or send to stdout
        f = openOutputFile(
            fileName, append=appendFile, delim=delim,
            fileCollisionMethod=fileCollisionMethod, encoding=encoding)

        # defer to pandas for actual data output. We're fetching a string
        # repr and then writeing to file ourselves
        # Include header line if not matrixOnly
        datStr = self.data.to_csv(sep=delim,
                                  columns=self.columns,  # sets the order
                                  header=(not matrixOnly),
                                  index=False)
        f.write(datStr)

        if f != sys.stdout:
            f.close()
            logging.info('saved wide-format data to %s' % f.name)

    def addData(self, thisType, value):
        """Add a piece of data to the current trial
        """
        # store in the columns list to help ordering later
        if thisType not in self.columns:
            self.columns.append(thisType)
        # save the actual value in a data dict
        self.thisTrial[thisType] = value
        if self.getExp() is not None:
            # update the experiment handler too
            self.getExp().addData(thisType, value)


class TrialHandlerExt(TrialHandler):
    """A class for handling trial sequences in a *non-counterbalanced design*
    (i.e. *oddball paradigms*). Its functions are a superset of the
    class TrialHandler, and as such, can also be used for normal trial
    handling.

    TrialHandlerExt has the same function names for data storage facilities.

    To use non-counterbalanced designs, all TrialType dict entries in the
    trial list must have a key called "weight". For example, if you want
    trial types A, B, C, and D to have 10, 5, 3, and 2 repetitions per
    block, then the trialList can look like:

    [{Name:'A', ..., weight:10},
     {Name:'B', ..., weight:5},
     {Name:'C', ..., weight:3},
     {Name:'D', ..., weight:2}]

    For experimenters using an excel or csv file for trial list, a column
    called weight is appropriate for this purpose.

    Calls to .next() will fetch the next trial object given to this handler,
    according to the method specified (random, sequential, fullRandom).
    Calls will raise a StopIteration error when all trials are exhausted.

    *Authored by Suddha Sourav at BPN, Uni Hamburg - heavily borrowing
    from the TrialHandler class*
    """

    def __init__(self,
                 trialList,
                 nReps,
                 method='random',
                 dataTypes=None,
                 extraInfo=None,
                 seed=None,
                 originPath=None,
                 name='',
                 autoLog=True):
        """

        :Parameters:

            trialList: a simple list (or flat array) of dictionaries
                specifying conditions. This can be imported from an
                excel / csv file using :func:`~psychopy.data.importConditions`
                For non-counterbalanced designs, each dict entry in
                trialList must have a key called weight!

            nReps: number of repeats for all conditions. When using a
                non-counterbalanced design, nReps is analogous to the number
                of blocks.

            method: *'random',* 'sequential', or 'fullRandom'
                When the weights are not specified:
                'sequential' presents the conditions in the order they appear
                in the list. 'random' will result in a shuffle of the
                conditions on each  repeat, but all conditions occur once
                before the second repeat etc. 'fullRandom' fully randomises
                the trials across repeats as well, which means you could
                potentially run all trials of one condition before any trial
                of another.

                In the presence of weights:
                'sequential' presents each trial type the number of times
                specified by its weight, before moving on to the next type.
                'random' randomizes the presentation order within block.
                'fulLRandom' shuffles trial order across weights an nRep,
                that is, a full shuffling.


            dataTypes: (optional) list of names for data storage. e.g.
                ['corr','rt','resp']. If not provided then these will be
                created as needed during calls to
                :func:`~psychopy.data.TrialHandler.addData`

            extraInfo: A dictionary
                This will be stored alongside the data and usually describes
                the experiment and subject ID, date etc.

            seed: an integer
                If provided then this fixes the random number generator
                to use the same pattern
                of trials, by seeding its startpoint

            originPath: a string describing the location of the script /
                experiment file path. The psydat file format will store a
                copy of the experiment if possible. If `originPath==None`
                is provided here then the TrialHandler will still store a
                copy of the script where it was created. If `OriginPath==-1`
                then nothing will be stored.

        :Attributes (after creation):

            .data - a dictionary of numpy arrays, one for each data type
                stored

            .trialList - the original list of dicts, specifying the conditions

            .thisIndex - the index of the current trial in the original
                conditions list

            .nTotal - the total number of trials that will be run

            .nRemaining - the total number of trials remaining

            .thisN - total trials completed so far

            .thisRepN - which repeat you are currently on

            .thisTrialN - which trial number *within* that repeat

            .thisTrial - a dictionary giving the parameters of the current
                trial

            .finished - True/False for have we finished yet

            .extraInfo - the dictionary of extra info as given at beginning

            .origin - the contents of the script or builder experiment that
                created the handler

            .trialWeights - None if all weights are not specified. If all
                weights are specified, then a list containing the weights
                of the trial types.

        """
        self.name = name
        self.autoLog = autoLog

        if trialList in (None, []):
            # user wants an empty trialList
            # which corresponds to a list with a single empty entry
            self.trialList = [None]
        # user has hopefully specified a filename
        elif isinstance(trialList, basestring) and os.path.isfile(trialList):
            # import conditions from that file
            self.trialList = importConditions(trialList)
        else:
            self.trialList = trialList
        # convert any entry in the TrialList into a TrialType object (with
        # obj.key or obj[key] access)
        for n, entry in enumerate(self.trialList):
            if type(entry) == dict:
                self.trialList[n] = TrialType(entry)
        self.nReps = nReps
        # Add Su
        if not trialList or not all('weight' in d for d in trialList):
            self.trialWeights = None
            self.nTotal = self.nReps * len(self.trialList)
        else:
            self.trialWeights = [d['weight'] for d in trialList]
            self.nTotal = self.nReps * sum(self.trialWeights)
        self.nRemaining = self.nTotal  # subtract 1 each trial
        self.method = method
        self.thisRepN = 0  # records which repetition or pass we are on
        self.thisTrialN = -1  # records trial number within this repetition
        self.thisN = -1
        self.thisIndex = 0  # index of current trial in the conditions list
        self.thisTrial = []
        self.finished = False
        self.extraInfo = extraInfo
        self.seed = seed
        # create dataHandler
        if self.trialWeights is None:
            self.data = DataHandler(trials=self)
        else:
            self.data = DataHandler(trials=self,
                                    dataShape=[sum(self.trialWeights), nReps])
        if dataTypes is not None:
            self.data.addDataType(dataTypes)
        self.data.addDataType('ran')
        self.data['ran'].mask = False  # bool - all entries are valid
        self.data.addDataType('order')
        # generate stimulus sequence
        if self.method in ('random', 'sequential', 'fullRandom'):
            self.sequenceIndices = self._createSequence()
        else:
            self.sequenceIndices = []

        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None  # the experiment handler that owns me!

    def _createSequence(self):
        """Pre-generates the sequence of trial presentations (for
        non-adaptive methods). This is called automatically when the
        TrialHandler is initialised so doesn't need an explicit call
        from the user.

        The returned sequence has form indices[stimN][repN]
        Example: sequential with 6 trialtypes (rows), 5 reps (cols), returns:
            [[0 0 0 0 0]
             [1 1 1 1 1]
             [2 2 2 2 2]
             [3 3 3 3 3]
             [4 4 4 4 4]
             [5 5 5 5 5]]
        These 30 trials will be returned by .next() in the order:
            0, 1, 2, 3, 4, 5,   0, 1, 2, ...  ... 3, 4, 5

        Example: random, with 3 trialtypes, where the weights of
        conditions 0,1, and 2 are 3,2, and 1 respectively,
        and a rep value of 5, might return:
            [[0 1 2 0 1]
             [1 0 1 1 1]
             [0 2 0 0 0]
             [0 0 0 1 0]
             [2 0 1 0 2]
             [1 1 0 2 0]]

        These 30 trials will be returned by .next() in the order:
            0, 1, 0, 0, 2, 1,   1, 0, 2, 0, 0, 1, ...
            ... 0, 2, 0  *stopIteration*

        To add a new type of sequence (as of v1.65.02):
        - add the sequence generation code here
        - adjust "if self.method in [ ...]:" in both __init__ and .next()
        - adjust allowedVals in experiment.py -> shows up in DlgLoopProperties
        Note that users can make any sequence whatsoever outside of PsychoPy,
        and specify sequential order; any order is possible this way.
        """
        # create indices for a single rep
        indices = numpy.asarray(self._makeIndices(self.trialList), dtype=int)

        repeat = numpy.repeat
        reshape = numpy.reshape
        if self.method == 'random':
            seqIndices = []
            seed = self.seed
            for thisRep in range(self.nReps):
                if self.trialWeights is None:
                    idx = indices.flat
                else:
                    idx = repeat(indices, self.trialWeights)
                thisRepSeq = shuffleArray(idx, seed=seed).tolist()
                seed = None  # so that we only seed the first pass through!
                seqIndices.append(thisRepSeq)
            seqIndices = numpy.transpose(seqIndices)
        elif self.method == 'sequential':
            if self.trialWeights is None:
                seqIndices = repeat(indices, self.nReps, 1)
            else:
                _base = repeat(indices, self.trialWeights, 0)
                seqIndices = repeat(_base, self.nReps, 1)
        elif self.method == 'fullRandom':
            if self.trialWeights is None:
                # indices * nReps, flatten, shuffle, unflatten;
                # only use seed once
                sequential = repeat(indices, self.nReps, 1)
                randomFlat = shuffleArray(sequential.flat, seed=self.seed)
                seqIndices = reshape(randomFlat,
                                     (len(indices), self.nReps))
            else:
                _base = repeat(indices, self.trialWeights, 0)
                sequential = repeat(_base, self.nReps, 1)
                randomFlat = shuffleArray(sequential.flat, seed=self.seed)
                seqIndices = reshape(randomFlat,
                                     (sum(self.trialWeights), self.nReps))

        if self.autoLog:
            # Change
            msg = 'Created sequence: %s, trialTypes=%d, nReps=%d, seed=%s'
            vals = (self.method, len(indices), self.nReps, str(self.seed))
            logging.exp(msg % vals)
        return seqIndices

    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; thisTrial, thisTrialN and thisIndex
        If the trials have ended this method will raise a StopIteration error.
        This can be handled with code such as::

            trials = data.TrialHandler(.......)
            for eachTrial in trials:  # automatically stops when done
                # do stuff

        or::

            trials = data.TrialHandler(.......)
            while True:  # ie forever
                try:
                    thisTrial = trials.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial
        """
        # update pointer for next trials
        self.thisTrialN += 1  # number of trial this pass
        self.thisN += 1  # number of trial in total
        self.nRemaining -= 1

        if self.trialWeights is None:
            if self.thisTrialN == len(self.trialList):
                # start a new repetition
                self.thisTrialN = 0
                self.thisRepN += 1
        else:
            if self.thisTrialN == sum(self.trialWeights):
                # start a new repetition
                self.thisTrialN = 0
                self.thisRepN += 1

        if self.thisRepN >= self.nReps:
            # all reps complete
            self.thisTrial = []
            self.finished = True

        if self.finished == True:
            self._terminate()

        # fetch the trial info
        if self.method in ('random', 'sequential', 'fullRandom'):
            if self.trialWeights is None:
                idx = self.sequenceIndices[self.thisTrialN]
                self.thisIndex = idx[self.thisRepN]
                self.thisTrial = self.trialList[self.thisIndex]
                self.data.add('ran', 1)
                self.data.add('order', self.thisN)
            else:
                idx = self.sequenceIndices[self.thisTrialN]
                self.thisIndex = idx[self.thisRepN]
                self.thisTrial = self.trialList[self.thisIndex]

                self.data.add('ran', 1,
                              position=self.getNextTrialPosInDataHandler())
                # The last call already adds a ran to this trial, so get the
                # current pos now
                self.data.add('order', self.thisN,
                              position=self.getCurrentTrialPosInDataHandler())

        if self.autoLog:
            msg = 'New trial (rep=%i, index=%i): %s'
            vals = (self.thisRepN, self.thisTrialN, self.thisTrial)
            logging.exp(msg % vals, obj=self.thisTrial)
        return self.thisTrial

    def getCurrentTrialPosInDataHandler(self):
        # if there's no trial weights, then the current position is simply
        # [trialIndex, nRepetition]
        if self.trialWeights is None:
            repN = sum(self['ran'][self.trials.thisIndex]) - 1
            position = [self.trials.thisIndex, repN]
        else:
            # if there are trial weights, the situation is slightly more
            # involved, because the same index can be repeated for a number
            # of times. If we had a sequential array, then the rows in
            # DataHandler for that trialIndex would be from
            # sum(trialWeights[begin:trialIndex]) to
            # sum(trialWeights[begin:trialIndex+1]).

            # if we haven't begun the experiment yet, then the last row
            # of the first column is used as the current position,
            # emulating what TrialHandler does. The following two lines
            # also prevents calculating garbage position values in case
            # the first row has a null weight
            if self.thisN < 0:
                return [0, -1]

            firstRowIndex = sum(self.trialWeights[:self.thisIndex])
            lastRowIndex = sum(self.trialWeights[:self.thisIndex + 1])

            # get the number of the trial presented by summing in ran for the
            # rows above and all columns
            nThisTrialPresented = numpy.sum(
                self.data['ran'][firstRowIndex:lastRowIndex, :])

            _tw = self.trialWeights[self.thisIndex]
            dataRowThisTrial = firstRowIndex + (nThisTrialPresented - 1) % _tw
            dataColThisTrial = int((nThisTrialPresented - 1) / _tw)

            position = [dataRowThisTrial, dataColThisTrial]

        return position

    def getNextTrialPosInDataHandler(self):
        # if there's no trial weights, then the current position is
        # simply [trialIndex, nRepetition]
        if self.trialWeights is None:
            repN = sum(self['ran'][self.trials.thisIndex])
            position = [self.trials.thisIndex, repN]
        else:
            # if there are trial weights, the situation is slightly more
            # involved, because the same index can be repeated for a
            # number of times. If we had a sequential array, then the
            # rows in DataHandler for that trialIndex would
            # be from sum(trialWeights[begin:trialIndex]) to
            # sum(trialWeights[begin:trialIndex+1]).

            firstRowIndex = sum(self.trialWeights[:self.thisIndex])
            lastRowIndex = sum(self.trialWeights[:self.thisIndex + 1])

            # get the number of the trial presented by summing in ran for the
            # rows above and all columns
            nThisTrialPresented = numpy.sum(
                self.data['ran'][firstRowIndex:lastRowIndex, :])

            _tw = self.trialWeights[self.thisIndex]
            dataRowThisTrial = firstRowIndex + nThisTrialPresented % _tw
            dataColThisTrial = int(nThisTrialPresented / _tw)

            position = [dataRowThisTrial, dataColThisTrial]

        return position

    def addData(self, thisType, value, position=None):
        """Add data for the current trial
        """

        if self.trialWeights is None:
            pos = None
        else:
            pos = self.getCurrentTrialPosInDataHandler()
        self.data.add(thisType, value, position=pos)
        # change this!
        if self.getExp() is not None:
            # update the experiment handler too:
            self.getExp().addData(thisType, value)

    def _createOutputArrayData(self, dataOut):
        """This just creates the dataOut part of the output matrix.
        It is called by _createOutputArray() which creates the header
        line and adds the stimOut columns
        """

        if self.trialWeights is not None:
            # remember to use other array instead of self.data
            _vals = numpy.arange(len(self.trialList))
            idx_data = numpy.repeat(_vals, self.trialWeights)

        # list of data headers
        dataHead = []
        # will store data that has been analyzed
        dataAnal = dict([])
        if type(dataOut) == str:
            # don't do list convert or we get a list of letters
            dataOut = [dataOut]
        elif type(dataOut) != list:
            dataOut = list(dataOut)

        # expand any 'all' dataTypes to the full list of available dataTypes
        allDataTypes = self.data.keys()
        # treat these separately later
        allDataTypes.remove('ran')
        # ready to go through standard data types
        dataOutNew = []
        for thisDataOut in dataOut:
            if thisDataOut == 'n':
                # n is really just the sum of the ran trials
                dataOutNew.append('ran_sum')
                continue  # no need to do more with this one
            # then break into dataType and analysis
            dataType, analType = string.rsplit(thisDataOut, '_', 1)
            if dataType == 'all':
                keyType = [key + "_" + analType for key in allDataTypes]
                dataOutNew.extend(keyType)
                if 'order_mean' in dataOutNew:
                    dataOutNew.remove('order_mean')
                if 'order_std' in dataOutNew:
                    dataOutNew.remove('order_std')
            else:
                dataOutNew.append(thisDataOut)
        dataOut = dataOutNew
        # sort so that all datatypes come together, rather than all analtypes
        dataOut.sort()

        # do the various analyses, keeping track of fails (e.g. mean of a
        # string)
        dataOutInvalid = []
        # add back special data types (n and order)
        if 'ran_sum' in dataOut:
            # move n to the first column
            dataOut.remove('ran_sum')
            dataOut.insert(0, 'ran_sum')
        if 'order_raw' in dataOut:
            # move order_raw to the second column
            dataOut.remove('order_raw')
            dataOut.append('order_raw')
        # do the necessary analysis on the data
        for thisDataOutN, thisDataOut in enumerate(dataOut):
            dataType, analType = string.rsplit(thisDataOut, '_', 1)
            if not dataType in self.data:
                # that analysis can't be done
                dataOutInvalid.append(thisDataOut)
                continue

            if self.trialWeights is None:
                thisData = self.data[dataType]
            else:
                resizedData = numpy.ma.masked_array(
                    numpy.zeros((len(self.trialList),
                                 max(self.trialWeights) * self.nReps)),
                    numpy.ones((len(self.trialList),
                                max(self.trialWeights) * self.nReps),
                               dtype=bool))
                for curTrialIndex in range(len(self.trialList)):
                    thisDataChunk = self.data[dataType][
                        idx_data == curTrialIndex, :]
                    padWidth = (max(self.trialWeights) * self.nReps -
                                numpy.prod(thisDataChunk.shape))
                    thisDataChunkRowPadded = numpy.pad(
                        thisDataChunk.transpose().flatten().data,
                        (0, padWidth), mode='constant',
                        constant_values=(0, 0))
                    thisDataChunkRowPaddedMask = numpy.pad(
                        thisDataChunk.transpose().flatten().mask,
                        (0, padWidth), mode='constant',
                        constant_values=(0, True))

                    thisDataChunkRow = numpy.ma.masked_array(
                        thisDataChunkRowPadded,
                        mask=thisDataChunkRowPaddedMask)
                    resizedData[curTrialIndex, :] = thisDataChunkRow

                thisData = resizedData

            # set the header
            dataHead.append(dataType + '_' + analType)
            # analyse thisData using numpy module
            if analType in dir(numpy):
                try:
                    # this will fail if we try to take mean of a string
                    if analType == 'std':
                        thisAnal = numpy.std(thisData, axis=1, ddof=0)
                        # normalise by N-1 instead. This should work by
                        # setting ddof=1 but doesn't as of 08/2010
                        # (because of using a masked array?)
                        N = thisData.shape[1]
                        if N == 1:
                            thisAnal *= 0  # prevent a divide-by-zero error
                        else:
                            sqrt = numpy.sqrt
                            thisAnal = thisAnal * sqrt(N) / sqrt(N - 1)
                    else:
                        thisAnal = eval("numpy.%s(thisData,1)" % analType)
                except Exception:
                    # that analysis doesn't work
                    dataHead.remove(dataType + '_' + analType)
                    dataOutInvalid.append(thisDataOut)
                    continue  # to next analysis
            elif analType == 'raw':
                thisAnal = thisData
            else:
                raise AttributeError('You can only use analyses from numpy')
            # add extra cols to header if necess
            if len(thisAnal.shape) > 1:
                for n in range(thisAnal.shape[1] - 1):
                    dataHead.append("")
            dataAnal[thisDataOut] = thisAnal

        # remove invalid analyses (e.g. average of a string)
        for invalidAnal in dataOutInvalid:
            dataOut.remove(invalidAnal)
        return dataOut, dataAnal, dataHead

    def saveAsWideText(self, fileName,
                       delim='\t',
                       matrixOnly=False,
                       appendFile=True):
        """Write a text file with the session, stimulus, and data values
        from each trial in chronological order.

        That is, unlike 'saveAsText' and 'saveAsExcel':
         - each row comprises information from only a single trial.
         - no summarizing is done (such as collapsing to produce mean and
           standard deviation values across trials).

        This 'wide' format, as expected by R for creating dataframes, and
        various other analysis programs, means that some information must
        be repeated on every row.

        In particular, if the trialHandler's 'extraInfo' exists, then each
        entry in there occurs in every row. In builder, this will include
        any entries in the 'Experiment info' field of the
        'Experiment settings' dialog. In Coder, this information can be set
        using something like::

            myTrialHandler.extraInfo = {'SubjID':'Joan Smith',
                                        'Group':'Control'}

        :Parameters:

            fileName:
                if extension is not specified, '.csv' will be appended if
                the delimiter is ',', else '.txt' will be appended.
                Can include path info.

            delim:
                allows the user to use a delimiter other than the default
                tab ("," is popular with file extension ".csv")

            matrixOnly:
                outputs the data with no header row.

            appendFile:
                will add this output to the end of the specified file if
                it already exists.

        """
        if self.thisTrialN < 1 and self.thisRepN < 1:
            # if both are < 1 we haven't started
            logging.info('TrialHandler.saveAsWideText called but no trials'
                         ' completed. Nothing saved')
            return -1

        # create the file or send to stdout
        if appendFile:
            writeFormat = 'a'
        else:
            writeFormat = 'w'  # will overwrite a file
        if fileName == 'stdout':
            f = sys.stdout
        elif fileName[-4:] in ('.dlm', '.DLM', '.tsv', '.TSV',
                               '.txt', '.TXT', '.csv', '.CSV'):
            f = codecs.open(fileName, writeFormat, encoding="utf-8")
        else:
            if delim == ',':
                f = codecs.open(fileName + '.csv',
                                writeFormat, encoding="utf-8")
            else:
                f = codecs.open(fileName + '.txt',
                                writeFormat, encoding="utf-8")

        # collect parameter names related to the stimuli:
        if self.trialList[0]:
            header = self.trialList[0].keys()
        else:
            header = []
        # and then add parameter names related to data (e.g. RT)
        header.extend(self.data.dataTypes)

        # loop through each trial, gathering the actual values:
        dataOut = []
        trialCount = 0
        # total number of trials = number of trialtypes * number of
        # repetitions:

        repsPerType = {}
        for rep in range(self.nReps):
            if self.trialWeights is None:
                nRows = len(self.trialList)
            else:
                nRows = sum(self.trialWeights)

            for trialN in range(nRows):
                # find out what trial type was on this trial
                trialTypeIndex = self.sequenceIndices[trialN, rep]
                # determine which repeat it is for this trial
                if trialTypeIndex not in repsPerType.keys():
                    repsPerType[trialTypeIndex] = 0
                else:
                    repsPerType[trialTypeIndex] += 1

                # create a dictionary representing each trial:
                # this is wide format, so we want fixed information (e.g.
                # subject ID, date, etc) repeated every line if it exists:
                if self.extraInfo != None:
                    nextEntry = self.extraInfo.copy()
                else:
                    nextEntry = {}

                # add a trial number so the original order of the data can
                # always be recovered if sorted during analysis:
                trialCount += 1
                nextEntry["TrialNumber"] = trialCount

                # what repeat are we on for this trial type?
                trep = repsPerType[trialTypeIndex]
                # collect the value from each trial of the vars in the header:
                tti = trialTypeIndex
                for prmName in header:
                    # the header includes both trial and data variables, so
                    # need to check before accessing:
                    if self.trialList[tti] and prmName in self.trialList[tti]:
                        nextEntry[prmName] = self.trialList[tti][prmName]
                    elif prmName in self.data:
                        if self.trialWeights is None:
                            nextEntry[prmName] = self.data[prmName][tti][trep]
                        else:
                            firstRowIndex = sum(self.trialWeights[:tti])
                            _tw = self.trialWeights[tti]
                            row = firstRowIndex + rep % _tw
                            col = int(rep / _tw)
                            nextEntry[prmName] = self.data[prmName][row][col]
                    else:
                        # allow a null value if this parameter wasn't
                        # explicitly stored on this trial:
                        nextEntry[prmName] = ''

                # store this trial's data
                dataOut.append(nextEntry)

        # get the extra 'wide' parameter names into the header line:
        header.insert(0, "TrialNumber")
        if self.extraInfo is not None:
            for key in self.extraInfo:
                header.insert(0, key)

        # write a header row:
        if not matrixOnly:
            f.write(delim.join(header) + '\n')
        # write the data matrix:
        for trial in dataOut:
            line = delim.join([unicode(trial[prm]) for prm in header])
            f.write(line + '\n')

        if f != sys.stdout:
            f.close()
            logging.info('saved wide-format data to %s' % f.name)


def importTrialTypes(fileName, returnFieldNames=False):
    """importTrialTypes is DEPRECATED (as of v1.70.00)
    Please use `importConditions` for identical functionality.
    """
    logging.warning("importTrialTypes is DEPRECATED (as of v1.70.00). "
                    "Please use `importConditions` for identical "
                    "functionality.")
    return importConditions(fileName, returnFieldNames)


def sliceFromString(sliceString):
    """Convert a text string into a valid slice object
    which can be used as indices for a list or array.

    >>> sliceFromString("0:10")
    slice(0,10,None)
    >>> sliceFromString("0::3")
    slice(0,None,3)
    >>> sliceFromString("-8:")
    slice(-8,None,None)
    """
    sliceArgs = []
    for val in sliceString.split(':'):
        if len(val) == 0:
            sliceArgs.append(None)
        else:
            sliceArgs.append(int(round(float(val))))
            # nb int(round(float(x))) is needed for x='4.3'
    return apply(slice, sliceArgs)


def indicesFromString(indsString):
    """Convert a text string into a valid list of indices
    """
    # "6"
    try:
        inds = int(round(float(indsString)))
        return [inds]
    except Exception:
        pass
    # "-6::2"
    try:
        inds = sliceFromString(indsString)
        return inds
    except Exception:
        pass
    # "1,4,8"
    try:
        inds = list(eval(indsString))
        return inds
    except Exception:
        pass


def importConditions(fileName, returnFieldNames=False, selection=""):
    """Imports a list of conditions from an .xlsx, .csv, or .pkl file

    The output is suitable as an input to :class:`TrialHandler`
    `trialTypes` or to :class:`MultiStairHandler` as a `conditions` list.

    If `fileName` ends with:

        - .csv:  import as a comma-separated-value file
            (header + row x col)
        - .xlsx: import as Excel 2007 (xlsx) files.
            No support for older (.xls) is planned.
        - .pkl:  import from a pickle file as list of lists
            (header + row x col)

    The file should contain one row per type of trial needed and one column
    for each parameter that defines the trial type. The first row should give
    parameter names, which should:

        - be unique
        - begin with a letter (upper or lower case)
        - contain no spaces or other punctuation (underscores are permitted)


    `selection` is used to select a subset of condition indices to be used
    It can be a list/array of indices, a python `slice` object or a string to
    be parsed as either option.
    e.g.:

        - "1,2,4" or [1,2,4] or (1,2,4) are the same
        - "2:5"       # 2, 3, 4 (doesn't include last whole value)
        - "-10:2:"    # tenth from last to the last in steps of 2
        - slice(-10, 2, None)  # the same as above
        - random(5) * 8  # five random vals 0-8

    """

    def _assertValidVarNames(fieldNames, fileName):
        """screens a list of names as candidate variable names. if all
        names are OK, return silently; else raise ImportError with msg
        """
        if not all(fieldNames):
            msg = ('Conditions file %s: Missing parameter name(s); '
                   'empty cell(s) in the first row?')
            raise ImportError(msg % fileName)
        for name in fieldNames:
            OK, msg = isValidVariableName(name)
            if not OK:
                # tailor message to importConditions
                msg = msg.replace('Variables', 'Parameters (column headers)')
                raise ImportError('Conditions file %s: %s%s"%s"' %
                                  (fileName, msg, os.linesep * 2, name))

    if fileName in ['None', 'none', None]:
        if returnFieldNames:
            return [], []
        return []
    if not os.path.isfile(fileName):
        msg = 'Conditions file not found: %s'
        raise ImportError(msg % os.path.abspath(fileName))

    def pandasToDictList(dataframe):
        """Convert a pandas dataframe to a list of dicts.
        This helper function is used by csv or excel imports via pandas
        """
        # convert the resulting dataframe to a numpy recarray
        trialsArr = dataframe.to_records(index=False)
        if trialsArr.shape == ():
            # convert 0-D to 1-D with one element:
            trialsArr = trialsArr[numpy.newaxis]
        fieldNames = trialsArr.dtype.names
        _assertValidVarNames(fieldNames, fileName)
        # convert the record array into a list of dicts
        trialList = []
        for trialN, trialType in enumerate(trialsArr):
            thisTrial = {}
            for fieldN, fieldName in enumerate(fieldNames):
                val = trialsArr[trialN][fieldN]

                if type(val) in [unicode, str]:
                    if val.startswith('[') and val.endswith(']'):
                        # val = eval('%s' %unicode(val.decode('utf8')))
                        val = eval(val)
                elif type(val) == numpy.string_:
                    val = unicode(val.decode('utf-8'))
                    # if it looks like a list, convert it:
                    if val.startswith('[') and val.endswith(']'):
                        # val = eval('%s' %unicode(val.decode('utf8')))
                        val = eval(val)
                elif numpy.isnan(val):  # if it is a numpy.nan, convert to None
                    val = None
                thisTrial[fieldName] = val
            trialList.append(thisTrial)
        return trialList, fieldNames

    if fileName.endswith('.csv'):
        with open(fileName, 'rU') as fileUniv:
            # use pandas reader, which can handle commas in fields, etc
            trialsArr = pandas.read_csv(fileUniv, encoding='utf-8')
            logging.debug("Read csv file with pandas: {}".format(fileName))
            trialList, fieldNames = pandasToDictList(trialsArr)

    elif fileName.endswith(('.xlsx','.xls')) and haveXlrd:
        trialsArr = pandas.read_excel(fileName)
        logging.debug("Read excel file with pandas: {}".format(fileName))
        trialList, fieldNames = pandasToDictList(trialsArr)

    elif fileName.endswith('.xlsx'):
        if not haveOpenpyxl:
            raise ImportError('openpyxl or xlrd is required for loading excel '
                              'files, but neither was found.')
        if openpyxl.__version__ < "1.8":  # data_only added in 1.8
            wb = load_workbook(filename=fileName)
        else:
            wb = load_workbook(filename=fileName, data_only=True)
        ws = wb.worksheets[0]
        logging.debug("Read excel file with openpyxl: {}".format(fileName))
        try:
            # in new openpyxl (2.3.4+) get_highest_xx is deprecated
            nCols = ws.max_column
            nRows = ws.max_row
        except:
            # version openpyxl 1.5.8 (in Standalone 1.80) needs this
            nCols = ws.get_highest_column()
            nRows = ws.get_highest_row()

        # get parameter names from the first row header
        fieldNames = []
        for colN in range(nCols):
            fieldName = ws.cell(_getExcelCellName(col=colN, row=0)).value
            fieldNames.append(fieldName)
        _assertValidVarNames(fieldNames, fileName)

        # loop trialTypes
        trialList = []
        for rowN in range(1, nRows):  # skip header first row
            thisTrial = {}
            for colN in range(nCols):
                val = ws.cell(_getExcelCellName(col=colN, row=rowN)).value
                # if it looks like a list or tuple, convert it
                if (type(val) in (unicode, str) and
                        (val.startswith('[') and val.endswith(']') or
                         val.startswith('(') and val.endswith(')'))):
                    val = eval(val)
                fieldName = fieldNames[colN]
                thisTrial[fieldName] = val
            trialList.append(thisTrial)

    elif fileName.endswith('.pkl'):
        f = open(fileName, 'rU')  # is U needed?
        try:
            trialsArr = cPickle.load(f)
        except Exception:
            raise ImportError('Could not open %s as conditions' % fileName)
        f.close()
        trialList = []
        fieldNames = trialsArr[0]  # header line first
        _assertValidVarNames(fieldNames, fileName)
        for row in trialsArr[1:]:
            thisTrial = {}
            for fieldN, fieldName in enumerate(fieldNames):
                # type is correct, being .pkl
                thisTrial[fieldName] = row[fieldN]
            trialList.append(thisTrial)
    else:
        raise IOError('Your conditions file should be an '
                      'xlsx, csv or pkl file')

    # if we have a selection then try to parse it
    if isinstance(selection, basestring) and len(selection) > 0:
        selection = indicesFromString(selection)
        if not isinstance(selection, slice):
            for n in selection:
                try:
                    assert n == int(n)
                except Exception:
                    raise TypeError("importConditions() was given some "
                                    "`indices` but could not parse them")
    # the selection might now be a slice or a series of indices
    if isinstance(selection, slice):
        trialList = trialList[selection]
    elif len(selection) > 0:
        allConds = trialList
        trialList = []
        for ii in selection:
            trialList.append(allConds[int(round(ii))])

    logging.exp('Imported %s as conditions, %d conditions, %d params' %
                (fileName, len(trialList), len(fieldNames)))
    if returnFieldNames:
        return (trialList, fieldNames)
    else:
        return trialList


def createFactorialTrialList(factors):
    """Create a trialList by entering a list of factors with names (keys)
    and levels (values) it will return a trialList in which all factors
    have been factorially combined (so for example if there are two factors
    with 3 and 5 levels the trialList will be a list of 3*5 = 15, each
    specifying the values for a given trial

    Usage::

        trialList = createFactorialTrialList(factors)

    :Parameters:

        factors : a dictionary with names (keys) and levels (values) of the
            factors

    Example::

        factors={"text": ["red", "green", "blue"],
                 "letterColor": ["red", "green"],
                 "size": [0, 1]}
        mytrials = createFactorialTrialList(factors)
    """

    # the first step is to place all the factorial combinations in a list of
    # lists
    tempListOfLists = [[]]
    for key in factors:
        # this takes the levels of each factor as a set of values
        # (a list) at a time
        alist = factors[key]
        tempList = []
        for value in alist:
            # now we loop over the values in a given list,
            # and add each value of the other lists
            for iterList in tempListOfLists:
                tempList.append(iterList + [key, value])
        tempListOfLists = tempList

    # this second step is so we can return a list in the format of trialList
    trialList = []
    for atrial in tempListOfLists:
        keys = atrial[0::2]  # the even elements are keys
        values = atrial[1::2]  # the odd elements are values
        atrialDict = {}
        for i in range(len(keys)):
            # this combines the key with the value
            atrialDict[keys[i]] = values[i]
        # append one trial at a time to the final trialList
        trialList.append(atrialDict)

    return trialList


class StairHandler(_BaseTrialHandler):
    """Class to handle smoothly the selection of the next trial
    and report current values etc.
    Calls to next() will fetch the next object given to this
    handler, according to the method specified.

    See ``Demos >> ExperimentalControl >> JND_staircase_exp.py``

    The staircase will terminate when *nTrials* AND *nReversals* have
    been exceeded. If *stepSizes* was an array and has been exceeded
    before nTrials is exceeded then the staircase will continue
    to reverse.

    *nUp* and *nDown* are always considered as 1 until the first reversal
    is reached. The values entered as arguments are then used.

    """

    def __init__(self,
                 startVal,
                 nReversals=None,
                 stepSizes=4,  # dB stepsize
                 nTrials=0,
                 nUp=1,
                 nDown=3,  # correct responses before stim goes down
                 extraInfo=None,
                 method='2AFC',
                 stepType='db',
                 minVal=None,
                 maxVal=None,
                 originPath=None,
                 name='',
                 autoLog=True,
                 **kwargs):
        """
        :Parameters:

            startVal:
                The initial value for the staircase.

            nReversals:
                The minimum number of reversals permitted.
                If `stepSizes` is a list, but the minimum number of
                reversals to perform, `nReversals`, is less than the
                length of this list, PsychoPy will automatically increase
                the minimum number of reversals and emit a warning.

            stepSizes:
                The size of steps as a single value or a list (or array).
                For a single value the step size is fixed. For an array or
                list the step size will progress to the next entry
                at each reversal.

            nTrials:
                The minimum number of trials to be conducted. If the
                staircase has not reached the required number of reversals
                then it will continue.

            nUp:
                The number of 'incorrect' (or 0) responses before the
                staircase level increases.

            nDown:
                The number of 'correct' (or 1) responses before the
                staircase level decreases.

            extraInfo:
                A dictionary (typically) that will be stored along with
                collected data using
                :func:`~psychopy.data.StairHandler.saveAsPickle` or
                :func:`~psychopy.data.StairHandler.saveAsText` methods.

            stepType:
                specifies whether each step will be a jump of the given
                size in 'db', 'log' or 'lin' units ('lin' means this
                intensity will be added/subtracted)

            method:
                Not used and may be deprecated in future releases.

            stepType: *'db'*, 'lin', 'log'
                The type of steps that should be taken each time. 'lin'
                will simply add or subtract that amount each step, 'db'
                and 'log' will step by a certain number of decibels or
                log units (note that this will prevent your value ever
                reaching zero or less)

            minVal: *None*, or a number
                The smallest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            maxVal: *None*, or a number
                The largest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            Additional keyword arguments will be ignored.

        :Notes:

        The additional keyword arguments `**kwargs` might for example be
        passed by the `MultiStairHandler`, which expects a `label` keyword
        for each staircase. These parameters are to be ignored by the
        StairHandler.

        """
        self.name = name
        self.startVal = startVal
        self.nUp = nUp
        self.nDown = nDown
        self.extraInfo = extraInfo
        self.method = method
        self.stepType = stepType

        try:
            self.stepSizes = list(stepSizes)
        except TypeError:
            # stepSizes is not array-like / iterable, i.e., a scalar.
            self.stepSizes = [stepSizes]

        self._variableStep = True if len(self.stepSizes) > 1 else False
        self.stepSizeCurrent = self.stepSizes[0]

        if nReversals is not None and len(self.stepSizes) > nReversals:
            logging.warn(
                "Increasing number of minimum required reversals to "
                "the number of step sizes (%i)." % len(self.stepSizes)
            )
            self.nReversals = len(self.stepSizes)
        else:
            self.nReversals = nReversals

        # to terminate the nTrials must be exceeded and either
        self.nTrials = nTrials
        self.finished = False
        self.thisTrialN = -1
        # a dict of lists where each should have the same length as the main
        # data:
        self.otherData = {}
        self.data = []
        self.intensities = []
        self.reversalPoints = []
        self.reversalIntensities = []
        # initially it goes down but on every step:
        self.currentDirection = 'start'
        # correct since last stim change (minus are incorrect):
        self.correctCounter = 0
        self._nextIntensity = self.startVal
        self.minVal = minVal
        self.maxVal = maxVal
        self.autoLog = autoLog
        # a flag for the 1-up 1-down initial rule:
        self.initialRule = 0

        # self.originPath and self.origin (the contents of the origin file)
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None  # the experiment handler that owns me!

    def __iter__(self):
        return self

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial

        This is essential to advance the staircase to a new intensity level!

        Supplying an `intensity` value here indicates that you did not use
        the recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        self.data.append(result)

        # if needed replace the existing intensity with this custom one
        if intensity != None:
            self.intensities.pop()
            self.intensities.append(intensity)

        # increment the counter of correct scores
        if result == 1:
            if len(self.data) > 1 and self.data[-2] == result:
                # increment if on a run
                self.correctCounter += 1
            else:
                # or reset
                self.correctCounter = 1
        else:
            if len(self.data) > 1 and self.data[-2] == result:
                # increment if on a run
                self.correctCounter -= 1
            else:
                # or reset
                self.correctCounter = -1

        # add the current data to experiment if poss
        if self.getExp() is not None:  # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)
        self.calculateNextIntensity()

    def addOtherData(self, dataName, value):
        """Add additional data to the handler, to be tracked alongside
        the result data but not affecting the value of the staircase
        """
        if not dataName in self.otherData:  # init the list
            if self.thisTrialN > 0:
                # might have run trals already
                self.otherData[dataName] = [None] * (self.thisTrialN - 1)
            else:
                self.otherData[dataName] = []
        # then add current value
        self.otherData[dataName].append(value)
        # add the current data to experiment if poss
        if self.getExp() != None:  # update the experiment handler too
            self.getExp().addData(dataName, value)

    def addData(self, result, intensity=None):
        """Deprecated since 1.79.00: This function name was ambiguous.
        Please use one of these instead:

            .addResponse(result, intensity)
            .addOtherData('dataName', value')

        """
        self.addResponse(result, intensity)

    def calculateNextIntensity(self):
        """Based on current intensity, counter of correct responses, and
        current direction.
        """

        if len(self.reversalIntensities) < 1:
            # always using a 1-down, 1-up rule initially
            if self.data[-1] == 1:  # last answer correct
                # got it right
                if self.currentDirection == 'up':
                    reversal = True
                else:
                    # direction is 'down' or 'start'
                    reversal = False
                self.currentDirection = 'down'
            else:
                # got it wrong
                if self.currentDirection == 'down':
                    reversal = True
                else:
                    # direction is 'up' or 'start'
                    reversal = False
                # now:
                self.currentDirection = 'up'
        elif self.correctCounter >= self.nDown:
            # n right, time to go down!
            if self.currentDirection != 'down':
                reversal = True
            else:
                reversal = False
            self.currentDirection = 'down'
        elif self.correctCounter <= -self.nUp:
            # n wrong, time to go up!
            # note current direction
            if self.currentDirection != 'up':
                reversal = True
            else:
                reversal = False
            self.currentDirection = 'up'
        else:
            # same as previous trial
            reversal = False

        # add reversal info
        if reversal:
            self.reversalPoints.append(self.thisTrialN)
            if len(self.reversalIntensities) < 1:
                self.initialRule = 1
            self.reversalIntensities.append(self.intensities[-1])

        # test if we're done
        if (len(self.reversalIntensities) >= self.nReversals and
                len(self.intensities) >= self.nTrials):
            self.finished = True
        # new step size if necessary
        if reversal and self._variableStep:
            if len(self.reversalIntensities) >= len(self.stepSizes):
                # we've gone beyond the list of step sizes
                # so just use the last one
                self.stepSizeCurrent = self.stepSizes[-1]
            else:
                _sz = len(self.reversalIntensities)
                self.stepSizeCurrent = self.stepSizes[_sz]

        # apply new step size
        if len(self.reversalIntensities) < 1 or self.initialRule == 1:
            self.initialRule = 0  # reset the flag
            if self.data[-1] == 1:
                self._intensityDec()
            else:
                self._intensityInc()
        elif self.correctCounter >= self.nDown:
            # n right, so going down
            self._intensityDec()
        elif self.correctCounter <= -self.nUp:
            # n wrong, so going up
            self._intensityInc()

    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; `thisTrial`, `thisTrialN` and `thisIndex`.

        If the trials have ended, calling this method will raise a
        StopIteration error. This can be handled with code such as::

            staircase = data.StairHandler(.......)
            for eachTrial in staircase:  # automatically stops when done
                # do stuff

        or::

            staircase = data.StairHandler(.......)
            while True:  # ie forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial

        """
        if self.finished == False:
            # check that all 'otherData' is aligned with current trialN
            for key in self.otherData.keys():
                while len(self.otherData[key]) < self.thisTrialN:
                    self.otherData[key].append(None)
            # update pointer for next trial
            self.thisTrialN += 1
            self.intensities.append(self._nextIntensity)
            return self._nextIntensity
        else:
            self._terminate()

    def _intensityInc(self):
        """increment the current intensity and reset counter
        """
        if self.stepType == 'db':
            self._nextIntensity *= 10.0**(self.stepSizeCurrent / 20.0)
        elif self.stepType == 'log':
            self._nextIntensity *= 10.0**self.stepSizeCurrent
        elif self.stepType == 'lin':
            self._nextIntensity += self.stepSizeCurrent
        # check we haven't gone out of the legal range
        if self._nextIntensity > self.maxVal and self.maxVal is not None:
            self._nextIntensity = self.maxVal
        self.correctCounter = 0

    def _intensityDec(self):
        """decrement the current intensity and reset counter
        """
        if self.stepType == 'db':
            self._nextIntensity /= 10.0**(self.stepSizeCurrent / 20.0)
        if self.stepType == 'log':
            self._nextIntensity /= 10.0**self.stepSizeCurrent
        elif self.stepType == 'lin':
            self._nextIntensity -= self.stepSizeCurrent
        self.correctCounter = 0
        # check we haven't gone out of the legal range
        if (self._nextIntensity < self.minVal) and self.minVal is not None:
            self._nextIntensity = self.minVal

    def saveAsText(self, fileName,
                   delim=None,
                   matrixOnly=False,
                   fileCollisionMethod='rename',
                   encoding='utf-8'):
        """Write a text file with the data

        :Parameters:

            fileName: a string
                The name of the file, including path if needed. The extension
                `.tsv` will be added if not included.

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted,
                ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided
                at initialisation.

            fileCollisionMethod:
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`

            encoding:
                The encoding to use when saving a the file.
                Defaults to `utf-8`.

        """

        if self.thisTrialN < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsText called but no '
                              'trials completed. Nothing saved')
            return -1

        # set default delimiter if none given
        if delim is None:
            delim = genDelimiter(fileName)

        # create the file or send to stdout
        f = openOutputFile(
            fileName, append=False, delim=delim,
            fileCollisionMethod=fileCollisionMethod, encoding=encoding)

        # write the data
        reversalStr = str(self.reversalIntensities)
        reversalStr = string.replace(reversalStr, ',', delim)
        reversalStr = string.replace(reversalStr, '[', '')
        reversalStr = string.replace(reversalStr, ']', '')
        f.write('\nreversalIntensities=\t%s\n' % reversalStr)

        reversalPts = str(self.reversalPoints)
        reversalPts = string.replace(reversalPts, ',', delim)
        reversalPts = string.replace(reversalPts, '[', '')
        reversalPts = string.replace(reversalPts, ']', '')
        f.write('reversalIndices=\t%s\n' % reversalPts)

        rawIntens = str(self.intensities)
        rawIntens = string.replace(rawIntens, ',', delim)
        rawIntens = string.replace(rawIntens, '[', '')
        rawIntens = string.replace(rawIntens, ']', '')
        f.write('\nintensities=\t%s\n' % rawIntens)

        responses = str(self.data)
        responses = string.replace(responses, ',', delim)
        responses = string.replace(responses, '[', '')
        responses = string.replace(responses, ']', '')
        f.write('responses=\t%s\n' % responses)

        # add self.extraInfo
        if self.extraInfo is not None and not matrixOnly:
            strInfo = str(self.extraInfo)
            # dict begins and ends with {} - remove
            # string.replace(strInfo, '{','')
            # strInfo = string.replace(strInfo, '}','')
            strInfo = strInfo[1:-1]
            # separate value from keyname
            strInfo = string.replace(strInfo, ': ', ':\n')
            # separate values from each other
            strInfo = string.replace(strInfo, ',', '\n')
            strInfo = string.replace(strInfo, 'array([ ', '')
            strInfo = string.replace(strInfo, '])', '')

            f.write('\n%s\n' % strInfo)

        f.write("\n")
        if f != sys.stdout:
            f.close()
            if self.autoLog:
                logging.info('saved data to %s' % f.name)

    def saveAsExcel(self, fileName, sheetName='data',
                    matrixOnly=False, appendFile=True,
                    fileCollisionMethod='rename'):
        """Save a summary data file in Excel OpenXML format workbook
        (:term:`xlsx`) for processing in most spreadsheet packages.
        This format is compatible with versions of Excel (2007 or greater)
        and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files
        (see :func:`TrialHandler.saveAsText()` ) that data can be stored
        in multiple named sheets within the file. So you could have a
        single file named after your experiment and then have one worksheet
        for each participant. Or you could have one file for each participant
        and then multiple sheets for repeated sessions etc.

        The file extension `.xlsx` will be added if not given already.

        The file will contain a set of values specifying the staircase level
        ('intensity') at each reversal, a list of reversal indices
        (trial numbers), the raw staircase / intensity level on *every*
        trial and the corresponding responses of the participant on every
        trial.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include
                relative or absolute path

            sheetName: string
                the name of the worksheet within the file

            matrixOnly: True or False
                If set to True then only the data itself will be output
                (no additional info)

            appendFile: True or False
                If False any existing file with this name will be
                overwritten. If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will
                be added to make it unique.

            fileCollisionMethod: string
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`
                This is ignored if ``append`` is ``True``.

        """

        if self.thisTrialN < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsExcel called but no '
                              'trials completed. Nothing saved')
            return -1
        # NB this was based on the limited documentation for openpyxl v1.0
        if not haveOpenpyxl:
            raise ImportError('openpyxl is required for saving files in '
                              'Excel (xlsx) format, but was not found.')
            # return -1

        # import necessary subpackages - they are small so won't matter to do
        # it here
        from openpyxl.workbook import Workbook
        from openpyxl.reader.excel import load_workbook

        if not fileName.endswith('.xlsx'):
            fileName += '.xlsx'
        # create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook = False
        else:
            if not appendFile:
                # the file exists but we're not appending, will be overwritten
                fileName = handleFileCollision(fileName,
                                               fileCollisionMethod)
            wb = Workbook()
            wb.properties.creator = 'PsychoPy' + psychopy.__version__
            newWorkbook = True

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title = sheetName
        else:
            ws = wb.create_sheet()
            ws.title = sheetName

        # write the data
        # reversals data
        ws.cell('A1').value = 'Reversal Intensities'
        ws.cell('B1').value = 'Reversal Indices'
        for revN, revIntens in enumerate(self.reversalIntensities):
            _cell = _getExcelCellName(col=0, row=revN + 1)  # col 0
            ws.cell(_cell).value = unicode(revIntens)
            _cell = _getExcelCellName(col=1, row=revN + 1)  # col 1
            ws.cell(_cell).value = unicode(self.reversalPoints[revN])

        # trials data
        ws.cell('C1').value = 'All Intensities'
        ws.cell('D1').value = 'All Responses'
        for intenN, intensity in enumerate(self.intensities):
            ws.cell(_getExcelCellName(col=2, row=intenN + 1)
                    ).value = unicode(intensity)
            ws.cell(_getExcelCellName(col=3, row=intenN + 1)
                    ).value = unicode(self.data[intenN])

        # add other data
        col = 4
        if self.otherData is not None:
            # for varName in self.otherData:
            for key, val in self.otherData.items():
                ws.cell(_getExcelCellName(col=col, row=0)
                        ).value = unicode(key)
                for oDatN in range(len(self.otherData[key])):
                    ws.cell(
                        _getExcelCellName(col=col, row=oDatN + 1)
                        ).value = unicode(self.otherData[key][oDatN])
                col += 1

        # add self.extraInfo
        if self.extraInfo is not None and not matrixOnly:
            ws.cell(_getExcelCellName(col=startingCol,
                                      row=0)).value = 'extraInfo'
            rowN = 1
            for key, val in self.extraInfo.items():
                _cell = _getExcelCellName(col=col, row=rowN)
                ws.cell(_cell).value = unicode(key) + u':'
                _cell = _getExcelCellName(col=col+1, row=rowN)
                ws.cell(_cell).value = unicode(val)
                rowN += 1


        wb.save(filename=fileName)
        if self.autoLog:
            logging.info('saved data to %s' % fileName)

    def saveAsPickle(self, fileName, fileCollisionMethod='rename'):
        """Basically just saves a copy of self (with data) to a pickle file.

        This can be reloaded if necess and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        """
        if self.thisTrialN < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsPickle called but no '
                              'trials completed. Nothing saved')
            return -1

        # otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName += '.psydat'

        f = openOutputFile(fileName, append=False,
                           fileCollisionMethod=fileCollisionMethod)
        cPickle.dump(self, f)
        f.close()
        logging.info('saved data to %s' % f.name)


class QuestHandler(StairHandler):
    """Class that implements the Quest algorithm for quick measurement of
    psychophysical thresholds.

    Uses Andrew Straw's `QUEST <http://www.visionegg.org/Quest>`_, which is a
    Python port of Denis Pelli's Matlab code.

    Measures threshold using a Weibull psychometric function. Currently, it is
    not possible to use a different psychometric function.

    Threshold 't' is measured on an abstract 'intensity' scale, which
    usually corresponds to log10 contrast.

    The Weibull psychometric function:

    _e = -10**(beta * (x2 + xThreshold))
    p2 = delta * gamma + (1-delta) * (1 - (1 - gamma) * exp(_e))

    **Example**::

        # setup display/window
        ...
        # create stimulus
        stimulus = visual.RadialStim(win=win, tex='sinXsin', size=1,
                                     pos=[0,0], units='deg')
        ...
        # create staircase object
        # trying to find out the point where subject's response is 50 / 50
        # if wanted to do a 2AFC then the defaults for pThreshold and gamma
        # are good
        staircase = data.QuestHandler(staircase._nextIntensity, 0.2,
            pThreshold=0.63, gamma=0.01,
            nTrials=20, minVal=0, maxVal=1)
        ...
        while thisContrast in staircase:
            # setup stimulus
            stimulus.setContrast(thisContrast)
            stimulus.draw()
            win.flip()
            core.wait(0.5)
            # get response
            ...
            # inform QUEST of the response, needed to calculate next level
            staircase.addResponse(thisResp)
        ...
        # can now access 1 of 3 suggested threshold levels
        staircase.mean()
        staircase.mode()
        staircase.quantile(0.5)  # gets the median

    """

    def __init__(self,
                 startVal,
                 startValSd,
                 pThreshold=0.82,
                 nTrials=None,
                 stopInterval=None,
                 method='quantile',
                 beta=3.5,
                 delta=0.01,
                 gamma=0.5,
                 grain=0.01,
                 range=None,
                 extraInfo=None,
                 minVal=None,
                 maxVal=None,
                 staircase=None,
                 originPath=None,
                 name='',
                 autoLog=True,
                 **kwargs):
        """
        Typical values for pThreshold are:
            * 0.82 which is equivalent to a 3 up 1 down standard staircase
            * 0.63 which is equivalent to a 1 up 1 down standard staircase
                (and might want gamma=0.01)

        The variable(s) nTrials and/or stopSd must be specified.

        `beta`, `delta`, and `gamma` are the parameters of the Weibull
        psychometric function.

        :Parameters:

            startVal:
                Prior threshold estimate or your initial guess threshold.

            startValSd:
                Standard deviation of your starting guess threshold.
                Be generous with the sd as QUEST will have trouble finding
                the true threshold if it's more than one sd from your
                initial guess.

            pThreshold
                Your threshold criterion expressed as probability of
                response==1. An intensity offset is introduced into the
                psychometric function so that the threshold (i.e.,
                the midpoint of the table) yields pThreshold.

            nTrials: *None* or a number
                The maximum number of trials to be conducted.

            stopInterval: *None* or a number
                The minimum 5-95% confidence interval required in the
                threshold estimate before stopping. If both this and
                nTrials is specified, whichever happens first will
                determine when Quest will stop.

            method: *'quantile'*, 'mean', 'mode'
                The method used to determine the next threshold to test.
                If you want to get a specific threshold level at the end
                of your staircasing, please use the quantile, mean, and
                mode methods directly.

            beta: *3.5* or a number
                Controls the steepness of the psychometric function.

            delta: *0.01* or a number
                The fraction of trials on which the observer presses blindly.

            gamma: *0.5* or a number
                The fraction of trials that will generate response 1 when
                intensity=-Inf.

            grain: *0.01* or a number
                The quantization of the internal table.

            range: *None*, or a number
                The intensity difference between the largest and smallest
                intensity that the internal table can store. This interval
                will be centered on the initial guess tGuess. QUEST assumes
                that intensities outside of this range have zero prior
                probability (i.e., they are impossible).

            extraInfo:
                A dictionary (typically) that will be stored along with
                collected data using
                :func:`~psychopy.data.StairHandler.saveAsPickle` or
                :func:`~psychopy.data.StairHandler.saveAsText` methods.

            minVal: *None*, or a number
                The smallest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            maxVal: *None*, or a number
                The largest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            staircase: *None* or StairHandler
                Can supply a staircase object with intensities and results.
                Might be useful to give the quest algorithm more information
                if you have it. You can also call the importData function
                directly.

            Additional keyword arguments will be ignored.

        :Notes:

        The additional keyword arguments `**kwargs` might for example be
        passed by the `MultiStairHandler`, which expects a `label` keyword
        for each staircase. These parameters are to be ignored by the
        StairHandler.

        """
        StairHandler.__init__(
            self, startVal, nTrials=nTrials, extraInfo=extraInfo,
            method=method, stepType='lin', minVal=minVal,
            maxVal=maxVal, name=name, autoLog=autoLog)

        self.stopInterval = stopInterval

        startVal = startVal
        startValSd = startValSd
        self._questNextIntensity = startVal

        # Create Quest object
        self._quest = QuestObject(
            startVal, startValSd, pThreshold, beta, delta, gamma,
            grain=grain, range=range)

        # Import any old staircase data
        if staircase is not None:
            self.importData(staircase.intensities, staircase.data)
        # store the origin file and its path
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None
        self.autoLog = autoLog

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial

        Supplying an `intensity` value here indicates that you did not use the
        recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        # Process user supplied intensity
        if intensity is None:
            intensity = self._questNextIntensity
        else:
            intensity = intensity
            # Update the intensity.
            #
            # During the first trial, self.intensities will be of length 0,
            # so pop() would not work.
            if len(self.intensities) != 0:
                self.intensities.pop()  # remove the auto-generated one
            self.intensities.append(intensity)
        # Update quest
        self._quest.update(intensity, result)
        # Update other things
        self.data.append(result)
        # add the current data to experiment if poss
        if self.getExp() != None:  # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)

        self._checkFinished()
        if not self.finished:
            self.calculateNextIntensity()

    def importData(self, intensities, results):
        """import some data which wasn't previously given to the quest
        algorithm
        """
        # NOT SURE ABOUT CLASS TO USE FOR RAISING ERROR
        if len(intensities) != len(results):
            raise AttributeError("length of intensities and results input "
                                 "must be the same")
        self.incTrials(len(intensities))
        for intensity, result in zip(intensities, results):
            try:
                self.next()
                self.addResponse(result, intensity)
            except StopIteration:
                # would get a stop iteration if stopInterval set
                pass    # TODO: might want to check if nTrials is still good

    def calculateNextIntensity(self):
        """based on current intensity and counter of correct responses
        """
        self._intensity()
        # Check we haven't gone out of the legal range
        if self._nextIntensity > self.maxVal and self.maxVal is not None:
            self._nextIntensity = self.maxVal
        elif self._nextIntensity < self.minVal and self.minVal is not None:
            self._nextIntensity = self.minVal
        self._questNextIntensity = self._nextIntensity

    def _intensity(self):
        """assigns the next intensity level"""
        if self.method == 'mean':
            self._questNextIntensity = self._quest.mean()
        elif self.method == 'mode':
            self._questNextIntensity = self._quest.mode()
        elif self.method == 'quantile':
            self._questNextIntensity = self._quest.quantile()
        # else: maybe raise an error
        self._nextIntensity = self._questNextIntensity

    def mean(self):
        """mean of Quest posterior pdf
        """
        return self._quest.mean()

    def sd(self):
        """standard deviation of Quest posterior pdf
        """
        return self._quest.sd()

    def mode(self):
        """mode of Quest posterior pdf
        """
        return self._quest.mode()[0]

    def quantile(self, p=None):
        """quantile of Quest posterior pdf
        """
        return self._quest.quantile(quantileOrder=p)

    def confInterval(self, getDifference=False):
        """
        Return estimate for the 5%--95% confidence interval (CI).

        :Parameters:

            getDifference (bool)
                If ``True``, return the width of the confidence interval
                (95% - 5% percentiles). If ``False``, return an NumPy array
                with estimates for the 5% and 95% boundaries.

        :Returns:

            scalar or array of length 2.
        """
        interval = [self.quantile(0.05), self.quantile(0.95)]
        if getDifference:
            return abs(interval[0] - interval[1])
        else:
            return interval

    def incTrials(self, nNewTrials):
        """increase maximum number of trials
        Updates attribute: `nTrials`
        """
        self.nTrials += nNewTrials

    def simulate(self, tActual):
        """returns a simulated user response to the next intensity level
        presented by Quest, need to supply the actual threshold level
        """
        # Current estimated intensity level
        if self.method == 'mean':
            tTest = self._quest.mean()
        elif self.method == 'mode':
            tTest = self._quest.mode()
        elif self.method == 'quantile':
            tTest = self._quest.quantile()
        return self._quest.simulate(tTest, tActual)

    def next(self):
        """Advances to next trial and returns it.
        Updates attributes; `thisTrial`, `thisTrialN`, `thisIndex`,
        `finished`, `intensities`

        If the trials have ended, calling this method will raise a
        StopIteration error. This can be handled with code such as::

            staircase = data.QuestHandler(.......)
            for eachTrial in staircase:  # automatically stops when done
                # do stuff

        or::

            staircase = data.QuestHandler(.......)
            while True:  # i.e. forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial
        """
        if self.finished == False:
            # update pointer for next trial
            self.thisTrialN += 1
            self.intensities.append(self._nextIntensity)
            return self._nextIntensity
        else:
            self._terminate()

    def _checkFinished(self):
        """checks if we are finished
        Updates attribute: `finished`
        """
        if self.nTrials is not None and len(self.intensities) >= self.nTrials:
            self.finished = True
        elif (self.stopInterval is not None and
                self.confInterval(True) < self.stopInterval):
            self.finished = True
        else:
            self.finished = False


class PsiHandler(StairHandler):
    """Handler to implement the "Psi" adaptive psychophysical method
    (Kontsevich & Tyler, 1999).

    This implementation assumes the form of the psychometric function
    to be a cumulative Gaussian. Psi estimates the two free parameters
    of the psychometric function, the location (alpha) and slope (beta),
    using Bayes' rule and grid approximation of the posterior distribution.
    It chooses stimuli to present by minimizing the entropy of this grid.
    Because this grid is represented internally as a 4-D array, one must
    choose the intensity, alpha, and beta ranges carefully so as to avoid
    a Memory Error. Maximum likelihood is used to estimate Lambda, the most
    likely location/slope pair. Because Psi estimates the entire
    psychometric function, any threshold defined on the function may be
    estimated once Lambda is determined.

    It is advised that Lambda estimates are examined after completion of
    the Psi procedure. If the estimated alpha or beta values equal your
    specified search bounds, then the search range most likely did not
    contain the true value. In this situation the procedure should be
    repeated with appropriately adjusted bounds.

    Because Psi is a Bayesian method, it can be initialized with a prior
    from existing research. A function to save the posterior over Lambda
    as a Numpy binary file is included.

    Kontsevich & Tyler (1999) specify their psychometric function in terms
    of d'. PsiHandler avoids this and treats all parameters with respect
    to stimulus intensity. Specifically, the forms of the psychometric
    function assumed for Yes/No and Two Alternative Forced Choice (2AFC)
    are, respectively:

    _normCdf = norm.cdf(x, mean=alpha, sd=beta)
    Y(x) = .5 * delta + (1 - delta) * _normCdf

    Y(x) = .5 * delta + (1 - delta) * (.5 + .5 * _normCdf)
    """

    def __init__(self,
                 nTrials,
                 intensRange, alphaRange, betaRange,
                 intensPrecision, alphaPrecision, betaPrecision,
                 delta,
                 stepType='lin',
                 expectedMin=0.5,
                 prior=None,
                 fromFile=False,
                 extraInfo=None,
                 name=''):
        """Initializes the handler and creates an internal Psi Object for
        grid approximation.

        :Parameters:

            nTrials (int)
                The number of trials to run.

            intensRange (list)
                Two element list containing the (inclusive) endpoints of
                the stimuli intensity range.

            alphaRange  (list)
                Two element list containing the (inclusive) endpoints of
                the alpha (location parameter) range.

            betaRange   (list)
                Two element list containing the (inclusive) endpoints of
                the beta (slope parameter) range.

            intensPrecision (float or int)
                If stepType == 'lin', this specifies the step size of the
                stimuli intensity range. If stepType == 'log', this specifies
                the number of steps in the stimuli intensity range.

            alphaPrecision  (float)
                The step size of the alpha (location parameter) range.

            betaPrecision   (float)
                The step size of the beta (slope parameter) range.

            delta   (float)
                The guess rate.

            stepType    (str)
                The type of steps to be used when constructing the stimuli
                intensity range. If 'lin' then evenly spaced steps are used.
                If 'log' then logarithmically spaced steps are used.
                Defaults to 'lin'.

            expectedMin  (float)
                The expected lower asymptote of the psychometric function
                (PMF).

                For a Yes/No task, the PMF usually extends across the
                interval [0, 1]; here, `expectedMin` should be set to `0`.

                For a 2-AFC task, the PMF spreads out across [0.5, 1.0].
                Therefore, `expectedMin` should be set to `0.5` in this
                case, and the 2-AFC psychometric function described above
                going to be is used.

                Currently, only Yes/No and 2-AFC designs are supported.

                Defaults to 0.5, or a 2-AFC task.

            prior   (numpy.ndarray or str)
                Optional prior distribution with which to initialize the
                Psi Object. This can either be a numpy.ndarray object or
                the path to a numpy binary file (.npy) containing the ndarray.

            fromFile    (str)
                Flag specifying whether prior is a file pathname or not.

            extraInfo   (dict)
                Optional dictionary object used in PsychoPy's built-in
                logging system.

            name    (str)
                Optional name for the PsiHandler used in PsychoPy's built-in
                logging system.

        :Raises:

            NotImplementedError
                If the supplied `minVal` parameter implies an experimental
                design other than Yes/No or 2-AFC.

        """
        if expectedMin not in [0, 0.5]:
            raise NotImplementedError(
                'Currently, only Yes/No and 2-AFC designs are '
                'supported. Please specify either `expectedMin=0` '
                '(Yes/No) or `expectedMin=0.5` (2-AFC).')

        StairHandler.__init__(
            self, startVal=None, nTrials=nTrials, extraInfo=extraInfo,
            stepType=stepType, minVal=intensRange[0],
            maxVal=intensRange[1], name=name
        )

        # Create Psi object
        if prior is not None and fromFile:
            try:
                prior = numpy.load(prior)
            except IOError:
                logging.warning("The specified pickle file could not be "
                                "read. Using a uniform prior instead.")
                prior = None

        twoAFC = True if expectedMin == 0.5 else False
        self._psi = PsiObject(
            intensRange, alphaRange, betaRange, intensPrecision,
            alphaPrecision, betaPrecision, delta=delta,
            stepType=stepType, TwoAFC=twoAFC, prior=prior)

        self._psi.update(None)

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial. Supplying an `intensity` value here
        indicates that you did not use the
        recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        self.data.append(result)

        # if needed replace the existing intensity with this custom one
        if intensity is not None:
            self.intensities.pop()
            self.intensities.append(intensity)
        # add the current data to experiment if possible
        if self.getExp() is not None:
            # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)
        self._psi.update(result)

    def next(self):
        """Advances to next trial and returns it.
        """
        self._checkFinished()
        if self.finished == False:
            # update pointer for next trial
            self.thisTrialN += 1
            self.intensities.append(self._psi.nextIntensity)
            return self._psi.nextIntensity
        else:
            self._terminate()

    def _checkFinished(self):
        """checks if we are finished
        Updates attribute: `finished`
        """
        if self.nTrials is not None and len(self.intensities) >= self.nTrials:
            self.finished = True
        else:
            self.finished = False

    def estimateLambda(self):
        """Returns a tuple of (location, slope)
        """
        return self._psi.estimateLambda()

    def estimateThreshold(self, thresh, lamb=None):
        """Returns an intensity estimate for the provided probability.

        The optional argument 'lamb' allows thresholds to be estimated
        without having to recompute the maximum likelihood lambda.
        """
        if lamb is not None:
            try:
                if len(lamb) != 2:
                    msg = ("Invalid user-specified lambda pair. A "
                           "new estimate of lambda will be computed.")
                    warnings.warn(msg, SyntaxWarning)
                    lamb = None
            except TypeError:
                msg = ("Invalid user-specified lambda pair. A new "
                       "estimate of lambda will be computed.")
                warnings.warn(msg, SyntaxWarning)
                lamb = None
        return self._psi.estimateThreshold(thresh, lamb)

    def savePosterior(self, fileName, fileCollisionMethod='rename'):
        """Saves the posterior array over probLambda as a pickle file
        with the specified name.

        :Parameters:
        fileCollisionMethod : string
            Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        """
        try:
            if os.path.exists(fileName):
                fileName = handleFileCollision(
                    fileName,
                    fileCollisionMethod=fileCollisionMethod
                )
            self._psi.savePosterior(fileName)
        except IOError:
            warnings.warn("An error occurred while trying to save the "
                          "posterior array. Continuing without saving...")


class MultiStairHandler(_BaseTrialHandler):

    def __init__(self, stairType='simple', method='random',
                 conditions=None, nTrials=50, originPath=None,
                 name='', autoLog=True):
        """A Handler to allow easy interleaved staircase procedures
        (simple or QUEST).

        Parameters for the staircases, as used by the relevant
        :class:`StairHandler` or
        :class:`QuestHandler` (e.g. the `startVal`, `minVal`, `maxVal`...)
        should be specified in the `conditions` list and may vary between
        each staircase. In particular, the conditions /must/ include the
        a `startVal` (because this is a required argument to the above
        handlers) a `label` to tag the staircase and a `startValSd`
        (only for QUEST staircases). Any parameters not specified in the
        conditions file will revert to the default for that individual
        handler.

        If you need to custom the behaviour further you may want to
        look at the recipe on :ref:`interleavedStairs`.

        :params:

            stairType: 'simple' or 'quest'
                Use a :class:`StairHandler` or :class:`QuestHandler`

            method: 'random' or 'sequential'
                The stairs are shuffled in each repeat but not randomised
                more than that (so you can't have 3 repeats of the same
                staircase in a row unless it's the only one still running)

            conditions: a list of dictionaries specifying conditions
                Can be used to control parameters for the different staicases.
                Can be imported from an Excel file using
                `psychopy.data.importConditions`
                MUST include keys providing, 'startVal', 'label' and
                'startValSd' (QUEST only).
                The 'label' will be used in data file saving so should
                be unique.
                See Example Usage below.

            nTrials=50
                Minimum trials to run (but may take more if the staircase
                hasn't also met its minimal reversals.
                See :class:`~psychopy.data.StairHandler`

        Example usage::

            conditions=[
                {'label':'low', 'startVal': 0.1, 'ori':45},
                {'label':'high','startVal': 0.8, 'ori':45},
                {'label':'low', 'startVal': 0.1, 'ori':90},
                {'label':'high','startVal': 0.8, 'ori':90},
                ]
            stairs = data.MultiStairHandler(conditions=conditions, nTrials=50)

            for thisIntensity, thisCondition in stairs:
                thisOri = thisCondition['ori']

                # do something with thisIntensity and thisOri

                stairs.addResponse(correctIncorrect)  # this is ESSENTIAL

            # save data as multiple formats
            stairs.saveDataAsExcel(fileName)  # easy to browse
            stairs.saveAsPickle(fileName)  # contains more info

        """
        self.name = name
        self.autoLog = autoLog
        self.type = stairType
        self.method = method  # 'random' or 'sequential'
        self.conditions = conditions
        self.nTrials = nTrials
        self.finished = False
        self.totalTrials = 0
        self._checkArguments()
        # create staircases
        self.staircases = []  # all staircases
        self.runningStaircases = []  # staircases that haven't finished yet
        self.thisPassRemaining = []  # staircases to run this pass
        self._createStairs()

        # fetch first staircase/value (without altering/advancing it)
        self._startNewPass()
        self.currentStaircase = self.thisPassRemaining[0]  # take the first
        # gets updated by self.addData()
        self._nextIntensity = self.currentStaircase._nextIntensity
        # store the origin file and its path
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None  # the experiment handler that owns me!

    def _checkArguments(self):
        # Did we get a `conditions` parameter, correctly formatted?
        if not isinstance(self.conditions, collections.Iterable):
            raise TypeError(
                '`conditions` parameter passed to MultiStairHandler '
                'should be a list, not a %s.' % type(self.conditions))

        c0 = self.conditions[0]
        if type(c0) != dict:
            raise TypeError(
                '`conditions` passed to MultiStairHandler should be a '
                'list of python dictionaries, not a list of %ss.' %
                type(c0))

        # Did `conditions` contain the things we need?
        params = c0.keys()
        if self.type not in ['simple', 'quest', 'QUEST']:
            raise ValueError(
                'MultiStairHandler `stairType` should be \'simple\', '
                '\'QUEST\' or \'quest\', not \'%s\'' % self.type)

        if 'startVal' not in params:
            raise AttributeError('MultiStairHandler needs a parameter called '
                                 '`startVal` in conditions')
        if 'label' not in params:
            raise AttributeError('MultiStairHandler needs a parameter called'
                                 ' `label` in conditions')
        if self.type in ['QUEST', 'quest'] and 'startValSd' not in params:
            raise AttributeError(
                'MultiStairHandler needs a parameter called '
                '`startValSd` in conditions for QUEST staircases.')

    def _createStairs(self):
        for condition in self.conditions:
            # We create a copy, because we are going to remove items from
            # this dictionary in this loop, but don't want these
            # changes to alter the originals in self.conditions.
            args = dict(condition)

            # If no individual `nTrials` parameter was supplied for this
            # staircase, use the `nTrials` that were passed to
            # the MultiStairHandler on instantiation.
            if 'nTrials' not in args:
                args['nTrials'] = self.nTrials

            if self.type == 'simple':
                startVal = args.pop('startVal')
                thisStair = StairHandler(startVal, **args)
            elif self.type in ['QUEST', 'quest']:
                startVal = args.pop('startVal')
                startValSd = args.pop('startValSd')
                thisStair = QuestHandler(startVal, startValSd, **args)

            # This isn't normally part of handler.
            thisStair.condition = condition

            # And finally, add it to the list.
            self.staircases.append(thisStair)
            self.runningStaircases.append(thisStair)

    def __iter__(self):
        return self

    def next(self):
        """Advances to next trial and returns it.

        This can be handled with code such as::

            staircase = data.MultiStairHandler(.......)
            for eachTrial in staircase:  # automatically stops when done
                # do stuff here for the trial

        or::

            staircase = data.MultiStairHandler(.......)
            while True:  # ie forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial

        """
        # create a new set for this pass if needed
        if (not hasattr(self, 'thisPassRemaining') or
                not self.thisPassRemaining):
            if self.runningStaircases:
                self._startNewPass()
            else:
                self.finished = True
                raise StopIteration

        # fetch next staircase/value
        self.currentStaircase = self.thisPassRemaining.pop(
            0)  # take the first and remove it
        # if staircase.next() not called, staircaseHandler would not
        # save the first intensity,
        # Error: miss align intensities and responses
        # gets updated by self.addResponse()
        self._nextIntensity = self.currentStaircase.next()

        # return value
        if not self.finished:
            # inform experiment of the condition (but not intensity,
            # that might be overridden by user)
            if self.getExp() != None:
                exp = self.getExp()
                stair = self.currentStaircase
                for key, value in stair.condition.items():
                    exp.addData("%s.%s" % (self.name, key), value)
                exp.addData(self.name + '.thisIndex',
                            self.conditions.index(stair.condition))
                exp.addData(self.name + '.thisRepN', stair.thisTrialN + 1)
                exp.addData(self.name + '.thisN', self.totalTrials)
                exp.addData(self.name + '.direction', stair.currentDirection)
                exp.addData(self.name + '.stepSize', stair.stepSizeCurrent)
                exp.addData(self.name + '.stepType', stair.stepType)
                exp.addData(self.name + '.intensity', self._nextIntensity)
            return self._nextIntensity, self.currentStaircase.condition
        else:
            raise StopIteration

    def _startNewPass(self):
        """Create a new iteration of the running staircases for this pass.

        This is not normally needed by the user - it gets called at __init__
        and every time that next() runs out of trials for this pass.
        """
        self.thisPassRemaining = copy.copy(self.runningStaircases)
        if self.method == 'random':
            numpy.random.shuffle(self.thisPassRemaining)

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial

        This is essential to advance the staircase to a new intensity level!
        """
        self.currentStaircase.addResponse(result, intensity)
        if self.currentStaircase.finished:
            self.runningStaircases.remove(self.currentStaircase)
        # add the current data to experiment if poss
        if self.getExp() != None:  # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)
        self.totalTrials += 1

    def addOtherData(self, name, value):
        """Add some data about the current trial that will not be used to
        control the staircase(s) such as reaction time data
        """
        self.currentStaircase.addOtherData(name, value)

    def addData(self, result, intensity=None):
        """Deprecated 1.79.00: It was ambiguous whether you were adding
        the response (0 or 1) or some other data concerning the trial so
        there is now a pair of explicit methods:

            addResponse(corr,intensity) #some data that alters the next
                trial value
            addOtherData('RT', reactionTime) #some other data that won't
                control staircase

        """
        self.addResponse(result, intensity)
        if type(result) in (str, unicode):
            raise TypeError("MultiStairHandler.addData should only receive "
                            "corr / incorr. Use .addOtherData('datName',val)")

    def saveAsPickle(self, fileName, fileCollisionMethod='rename'):
        """Saves a copy of self (with data) to a pickle file.

        This can be reloaded later and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        """
        if self.totalTrials < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsPickle called but no '
                              'trials completed. Nothing saved')
            return -1

        # otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName += '.psydat'

        f = openOutputFile(fileName, append=False,
                           fileCollisionMethod=fileCollisionMethod)
        cPickle.dump(self, f)
        f.close()
        logging.info('saved data to %s' % f.name)

    def saveAsExcel(self, fileName, matrixOnly=False, appendFile=False,
                    fileCollisionMethod='rename'):
        """Save a summary data file in Excel OpenXML format workbook
        (:term:`xlsx`) for processing in most spreadsheet packages.
        This format is compatible with versions of Excel (2007 or greater)
        and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files (see
        :func:`TrialHandler.saveAsText()` )
        that the data from each staircase will be save in the same file, with
        the sheet name coming from the 'label' given in the dictionary of
        conditions during initialisation of the Handler.

        The file extension `.xlsx` will be added if not given already.

        The file will contain a set of values specifying the staircase level
        ('intensity') at each reversal, a list of reversal indices
        (trial numbers), the raw staircase/intensity level on *every* trial
        and the corresponding responses of the participant on every trial.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include
                relative or absolute path

            matrixOnly: True or False
                If set to True then only the data itself will be output
                (no additional info)

            appendFile: True or False
                If False any existing file with this name will be overwritten.
                If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will
                be added to make it unique.

            fileCollisionMethod: string
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`
                This is ignored if ``append`` is ``True``.

        """
        if self.totalTrials < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsExcel called but no'
                              ' trials completed. Nothing saved')
            return -1

        append = appendFile
        for thisStair in self.staircases:
            # make a filename
            label = thisStair.condition['label']
            thisStair.saveAsExcel(
                fileName, sheetName=label, matrixOnly=matrixOnly,
                appendFile=append, fileCollisionMethod=fileCollisionMethod)
            append = True

    def saveAsText(self, fileName,
                   delim=None,
                   matrixOnly=False,
                   fileCollisionMethod='rename',
                   encoding='utf-8'):
        """Write out text files with the data.

        For MultiStairHandler this will output one file for each staircase
        that was run, with _label added to the fileName that you specify above
        (label comes from the condition dictionary you specified when you
        created the Handler).

        :Parameters:

            fileName: a string
                The name of the file, including path if needed. The extension
                `.tsv` will be added if not included.

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted,
                ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided
                at initialisation.

            fileCollisionMethod:
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`

            encoding:
                The encoding to use when saving a the file.
                Defaults to `utf-8`.

        """
        if self.totalTrials < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsText called but no trials'
                              ' completed. Nothing saved')
            return -1
        for thisStair in self.staircases:
            # make a filename
            label = thisStair.condition['label']
            thisFileName = fileName + "_" + label
            thisStair.saveAsText(
                fileName=thisFileName, delim=delim, matrixOnly=matrixOnly,
                fileCollisionMethod=fileCollisionMethod, encoding=encoding
            )

    def printAsText(self,
                    delim='\t',
                    matrixOnly=False):
        """Write the data to the standard output stream

        :Parameters:

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted,
                ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided
                at initialisation.
        """
        nStairs = len(self.staircases)
        for stairN, thisStair in enumerate(self.staircases):
            if stairN < nStairs - 1:
                thisMatrixOnly = True  # no header info for first files
            else:
                thisMatrixOnly = matrixOnly
            # make a filename
            label = thisStair.condition['label']
            thisStair.saveAsText(fileName='stdout', delim=delim,
                                 matrixOnly=thisMatrixOnly)


class DataHandler(dict):
    """For handling data (used by TrialHandler, principally, rather than
    by users directly)

    Numeric data are stored as numpy masked arrays where the mask is set
    True for missing entries. When any non-numeric data (string, list or
    array) get inserted using DataHandler.add(val) the array is converted
    to a standard (not masked) numpy array with dtype='O' and where missing
    entries have value = "--".

    Attributes:
        - ['key']=data arrays containing values for that key
            (e.g. data['accuracy']=...)
        - dataShape=shape of data (x,y,...z,nReps)
        - dataTypes=list of keys as strings

    """

    def __init__(self, dataTypes=None, trials=None, dataShape=None):
        self.trials = trials
        self.dataTypes = []  # names will be added during addDataType
        self.isNumeric = {}
        # if given dataShape use it - otherwise guess!
        if dataShape:
            self.dataShape = dataShape
        elif self.trials:
            self.dataShape = list(numpy.asarray(trials.trialList, 'O').shape)
            self.dataShape.append(trials.nReps)

        # initialise arrays now if poss
        if dataTypes and self.dataShape:
            for thisType in dataTypes:
                self.addDataType(thisType)

    def addDataType(self, names, shape=None):
        """Add a new key to the data dictionary of particular shape if
        specified (otherwise the shape of the trial matrix in the trial
        handler. Data are initialised to be zero everywhere. Not needed
        by user: appropriate types will be added during initialisation
        and as each xtra type is needed.
        """
        if not shape:
            shape = self.dataShape
        if not isinstance(names, basestring):
            # recursively call this function until we have a string
            for thisName in names:
                self.addDataType(thisName)
        else:
            # create the appropriate array in the dict
            # initially use numpy masked array of floats with mask=True
            # for missing vals. convert to a numpy array with dtype='O'
            # if non-numeric data given. NB don't use masked array with
            # dytpe='O' together - they don't unpickle
            self[names] = numpy.ma.zeros(shape, 'f')  # masked array of floats
            self[names].mask = True
            # add the name to the list
            self.dataTypes.append(names)
            self.isNumeric[names] = True  # until we need otherwise

    def add(self, thisType, value, position=None):
        """Add data to an existing data type (and add a new one if necess)
        """
        if not thisType in self:
            self.addDataType(thisType)
        if position is None:
            # 'ran' is always the first thing to update
            repN = sum(self['ran'][self.trials.thisIndex])
            if thisType != 'ran':
                # because it has already been updated
                repN -= 1
            # make a list where 1st digit is trial number
            position = [self.trials.thisIndex]
            position.append(repN)

        # check whether data falls within bounds
        posArr = numpy.asarray(position)
        shapeArr = numpy.asarray(self.dataShape)
        if not numpy.alltrue(posArr < shapeArr):
            # array isn't big enough
            logging.warning('need a bigger array for: ' + thisType)
            # not implemented yet!
            self[thisType] = extendArr(self[thisType], posArr)
        # check for ndarrays with more than one value and for non-numeric data
        if (self.isNumeric[thisType] and
                ((type(value) == numpy.ndarray and len(value) > 1) or
                 (type(value) not in [float, int]))):
            self._convertToObjectArray(thisType)
        # insert the value
        self[thisType][position[0], int(position[1])] = value

    def _convertToObjectArray(self, thisType):
        """Convert this datatype from masked numeric array to unmasked
        object array
        """
        dat = self[thisType]
        # create an array of Object type
        self[thisType] = numpy.array(dat.data, dtype='O')
        # masked vals should be "--", others keep data
        # we have to repeat forcing to 'O' or text gets truncated to 4chars
        self[thisType] = numpy.where(dat.mask, '--', dat).astype('O')
        self.isNumeric[thisType] = False


class FitFunction(object):
    """Deprecated: - use the specific functions; FitWeibull, FitLogistic...
    """

    def __init__(self, *args, **kwargs):
        raise DeprecationWarning("FitFunction is now fully DEPRECATED: use"
                                 " FitLogistic, FitWeibull etc instead")


class _baseFunctionFit(object):
    """Not needed by most users except as a superclass for developing
    your own functions

    Derived classes must have _eval and _inverse methods with @staticmethods
    """

    def __init__(self, xx, yy, sems=1.0, guess=None, display=1,
                 expectedMin=0.5):
        super(_baseFunctionFit, self).__init__()
        self.xx = numpy.asarray(xx)
        self.yy = numpy.asarray(yy)
        self.sems = numpy.asarray(sems)
        self.expectedMin = expectedMin
        self.guess = guess
        # for holding error calculations:
        self.ssq = 0
        self.rms = 0
        self.chi = 0
        # do the calculations:
        self._doFit()

    def _doFit(self):
        """The Fit class that derives this needs to specify its _evalFunction
        """
        # get some useful variables to help choose starting fit vals
        # self.params = optimize.fmin_powell(self._getErr, self.params,
        #    (self.xx,self.yy,self.sems),disp=self.display)
        # self.params = optimize.fmin_bfgs(self._getErr, self.params, None,
        #    (self.xx,self.yy,self.sems),disp=self.display)
        global _chance
        _chance = self.expectedMin
        self.params, self.covar = optimize.curve_fit(
            self._eval, self.xx, self.yy, p0=self.guess, sigma=self.sems)
        self.ssq = self._getErr(self.params, self.xx, self.yy, 1.0)
        self.chi = self._getErr(self.params, self.xx, self.yy, self.sems)
        self.rms = self.ssq / len(self.xx)

    def _getErr(self, params, xx, yy, sems):
        mod = self.eval(xx, params)
        err = sum((yy - mod)**2 / sems)
        return err

    def eval(self, xx, params=None):
        """Evaluate xx for the current parameters of the model, or for
        arbitrary params if these are given.
        """
        if params is None:
            params = self.params
        global _chance
        _chance = self.expectedMin
        #_eval is a static method - must be done this way because the
        # curve_fit function doesn't want to have any `self` object as
        # first arg
        yy = self._eval(xx, *params)
        return yy

    def inverse(self, yy, params=None):
        """Evaluate yy for the current parameters of the model,
        or for arbitrary params if these are given.
        """
        if params is None:
            # so the user can set params for this particular inv
            params = self.params
        xx = self._inverse(yy, *params)
        return xx


class FitWeibull(_baseFunctionFit):
    """Fit a Weibull function (either 2AFC or YN)
    of the form::

        y = chance + (1.0-chance)*(1-exp( -(xx/alpha)**(beta) ))

    and with inverse::

        x = alpha * (-log((1.0-y)/(1-chance)))**(1.0/beta)

    After fitting the function you can evaluate an array of x-values
    with ``fit.eval(x)``, retrieve the inverse of the function with
    ``fit.inverse(y)`` or retrieve the parameters from ``fit.params``
    (a list with ``[alpha, beta]``)
    """
    # static methods have no `self` and this is important for
    # optimise.curve_fit
    @staticmethod
    def _eval(xx, alpha, beta):
        global _chance
        xx = numpy.asarray(xx)
        yy = _chance + (1.0 - _chance) * (1 - numpy.exp(-(xx / alpha)**beta))
        return yy

    @staticmethod
    def _inverse(yy, alpha, beta):
        global _chance
        xx = alpha * (-numpy.log((1.0 - yy) / (1 - _chance))) ** (1.0 / beta)
        return xx


class FitNakaRushton(_baseFunctionFit):
    """Fit a Naka-Rushton function
    of the form::

        yy = rMin + (rMax-rMin) * xx**n/(xx**n+c50**n)

    After fitting the function you can evaluate an array of x-values
    with ``fit.eval(x)``, retrieve the inverse of the function with
    ``fit.inverse(y)`` or retrieve the parameters from ``fit.params``
    (a list with ``[rMin, rMax, c50, n]``)

    Note that this differs from most of the other functions in
    not using a value for the expected minimum. Rather, it fits this
    as one of the parameters of the model."""
    # static methods have no `self` and this is important for
    # optimise.curve_fit
    @staticmethod
    def _eval(xx, c50, n, rMin, rMax):
        xx = numpy.asarray(xx)
        if c50 <= 0:
            c50 = 0.001
        if n <= 0:
            n = 0.001
        if rMax <= 0:
            n = 0.001
        if rMin <= 0:
            n = 0.001
        yy = rMin + (rMax - rMin) * (xx**n / (xx**n + c50**n))
        return yy

    @staticmethod
    def _inverse(yy, c50, n, rMin, rMax):
        yScaled = (yy - rMin) / (rMax - rMin)  # remove baseline and scale
        # do we need to shift while fitting?
        yScaled[yScaled < 0] = 0
        xx = (yScaled * c50**n / (1 - yScaled))**(1 / n)
        return xx


class FitLogistic(_baseFunctionFit):
    """Fit a Logistic function (either 2AFC or YN)
    of the form::

        y = chance + (1-chance)/(1+exp((PSE-xx)*JND))

    and with inverse::

        x = PSE - log((1-chance)/(yy-chance) - 1)/JND

    After fitting the function you can evaluate an array of x-values
    with ``fit.eval(x)``, retrieve the inverse of the function with
    ``fit.inverse(y)`` or retrieve the parameters from ``fit.params``
    (a list with ``[PSE, JND]``)
    """
    # static methods have no `self` and this is important for
    # optimise.curve_fit
    @staticmethod
    def _eval(xx, PSE, JND):
        global _chance
        chance = _chance
        xx = numpy.asarray(xx)
        yy = chance + (1 - chance) / (1 + numpy.exp((PSE - xx) * JND))
        return yy

    @staticmethod
    def _inverse(yy, PSE, JND):
        global _chance
        yy = numpy.asarray(yy)
        xx = PSE - numpy.log((1 - _chance) / (yy - _chance) - 1) / JND
        return xx


class FitCumNormal(_baseFunctionFit):
    """Fit a Cumulative Normal function (aka error function or erf)
    of the form::

        y = chance + (1-chance)*((special.erf((xx-xShift)/(sqrt(2)*sd))+1)*0.5)

    and with inverse::

        x = xShift+sqrt(2)*sd*(erfinv(((yy-chance)/(1-chance)-.5)*2))

    After fitting the function you can evaluate an array of x-values
    with fit.eval(x), retrieve the inverse of the function with
    fit.inverse(y) or retrieve the parameters from fit.params (a list
    with [centre, sd] for the Gaussian distribution forming the cumulative)

    NB: Prior to version 1.74 the parameters had different meaning, relating
    to xShift and slope of the function (similar to 1/sd). Although that is
    more in with the parameters for the Weibull fit, for instance, it is less
    in keeping with standard expectations of normal (Gaussian distributions)
    so in version 1.74.00 the parameters became the [centre,sd] of the normal
    distribution.

    """
    # static methods have no `self` and this is important for
    # optimise.curve_fit
    @staticmethod
    def _eval(xx, xShift, sd):
        global _chance
        xx = numpy.asarray(xx)
        # NB numpy.special.erf() goes from -1:1
        yy = (_chance + (1 - _chance) *
              ((special.erf((xx - xShift) / (numpy.sqrt(2) * sd)) + 1) * 0.5))
        return yy

    @staticmethod
    def _inverse(yy, xShift, sd):
        global _chance
        yy = numpy.asarray(yy)
        # xx = (special.erfinv((yy-chance)/(1-chance)*2.0-1)+xShift)/xScale
        # NB: numpy.special.erfinv() goes from -1:1
        xx = (xShift + numpy.sqrt(2) * sd *
              special.erfinv(((yy - _chance) / (1 - _chance) - 0.5) * 2))
        return xx

######################### End psychopy.data classes #########################


def bootStraps(dat, n=1):
    """Create a list of n bootstrapped resamples of the data

    SLOW IMPLEMENTATION (Python for-loop)

    Usage:
        ``out = bootStraps(dat, n=1)``

    Where:
        dat
            an NxM or 1xN array (each row is a different condition, each
            column is a different trial)
        n
            number of bootstrapped resamples to create

        out
            - dim[0]=conditions
            - dim[1]=trials
            - dim[2]=resamples
    """
    dat = numpy.asarray(dat)
    if len(dat.shape) == 1:
        # have presumably been given a series of data for one stimulus
        # adds a dimension (arraynow has shape (1,Ntrials))
        dat = numpy.array([dat])

    nTrials = dat.shape[1]
    # initialise a matrix to store output
    resamples = numpy.zeros(dat.shape + (n,), dat.dtype)
    rand = numpy.random.rand
    for stimulusN in range(dat.shape[0]):
        thisStim = dat[stimulusN, :]  # fetch data for this stimulus
        for sampleN in range(n):
            indices = numpy.floor(nTrials * rand(nTrials)).astype('i')
            resamples[stimulusN, :, sampleN] = numpy.take(thisStim, indices)
    return resamples


def functionFromStaircase(intensities, responses, bins=10):
    """Create a psychometric function by binning data from a staircase
    procedure. Although the default is 10 bins Jon now always uses 'unique'
    bins (fewer bins looks pretty but leads to errors in slope estimation)

    usage::

        intensity, meanCorrect, n = functionFromStaircase(intensities,
                                                          responses, bins)

    where:
            intensities
                are a list (or array) of intensities to be binned

            responses
                are a list of 0,1 each corresponding to the equivalent
                intensity value

            bins
                can be an integer (giving that number of bins) or 'unique'
                (each bin is made from aa data for exactly one intensity
                value)

            intensity
                a numpy array of intensity values (where each is the center
                of an intensity bin)

            meanCorrect
                a numpy array of mean % correct in each bin

            n
                a numpy array of number of responses contributing to each mean
    """
    # convert to arrays
    try:
        # concatenate if multidimensional
        intensities = numpy.concatenate(intensities)
        responses = numpy.concatenate(responses)
    except Exception:
        intensities = numpy.array(intensities)
        responses = numpy.array(responses)

    # sort the responses
    sort_ii = numpy.argsort(intensities)
    sortedInten = numpy.take(intensities, sort_ii)
    sortedResp = numpy.take(responses, sort_ii)

    binnedResp = []
    binnedInten = []
    nPoints = []
    if bins == 'unique':
        intensities = numpy.round(intensities, decimals=8)
        uniqueIntens = numpy.unique(intensities)
        for thisInten in uniqueIntens:
            theseResps = responses[intensities == thisInten]
            binnedInten.append(thisInten)
            binnedResp.append(numpy.mean(theseResps))
            nPoints.append(len(theseResps))
    else:
        pointsPerBin = len(intensities) / float(bins)
        for binN in range(bins):
            start = int(round(binN * pointsPerBin))
            stop = int(round((binN + 1) * pointsPerBin))
            thisResp = sortedResp[start:stop]
            thisInten = sortedInten[start:stop]
            binnedResp.append(numpy.mean(thisResp))
            binnedInten.append(numpy.mean(thisInten))
            nPoints.append(len(thisInten))

    return binnedInten, binnedResp, nPoints


def getDateStr(format="%Y_%b_%d_%H%M"):
    """Uses ``time.strftime()``_ to generate a string of the form
    2012_Apr_19_1531 for 19th April 3.31pm, 2012.
    This is often useful appended to data filenames to provide unique names.
    To include the year: getDateStr(format="%Y_%b_%d_%H%M")
    returns '2011_Mar_16_1307' depending on locale, can have unicode chars
    in month names, so utf_8_decode them
    For date in the format of the current localization, do:
        data.getDateStr(format=locale.nl_langinfo(locale.D_T_FMT))
    """
    now = time.strftime(format, time.localtime())
    try:
        now_decoded = codecs.utf_8_decode(now)[0]
    except UnicodeDecodeError:
        # '2011_03_16_1307'
        now_decoded = time.strftime("%Y_%m_%d_%H%M", time.localtime())

    return now_decoded


def checkValidFilePath(filepath, makeValid=True):
    """Checks whether file path location (e.g. is a valid folder)

    This should also check whether we have write-permissions to the folder
    but doesn't currently do that!

    added in: 1.90.00
    """
    folder = os.path.split(os.path.abspath(filepath))[0]
    if not os.path.isdir(folder):
        os.makedirs(folder)  # spit an error if we fail
    return True


def isValidVariableName(name):
    """Checks whether a certain string could be used as a valid variable.

    Usage::

        OK, msg = isValidVariableName(name)

    >>> isValidVariableName('name')
    (True, '')
    >>> isValidVariableName('0name')
    (False, 'Variables cannot begin with numeric character')
    >>> isValidVariableName('first second')
    (False, 'Variables cannot contain punctuation or spaces')
    >>> isValidVariableName('')
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(None)
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(23)
    (False, "Variables must be string-like")
    >>> isValidVariableName('a_b_c')
    (True, '')
    """
    if not name:
        return False, "Variables cannot be missing, None, or ''"
    if not type(name) in (str, unicode, numpy.string_, numpy.unicode_):
        return False, "Variables must be string-like"
    try:
        name = str(name)  # convert from unicode if possible
    except Exception:
        if type(name) in [unicode, numpy.unicode_]:
            msg = ("name %s (type %s) contains non-ASCII characters"
                   " (e.g. accents)")
            raise AttributeError(msg % (name, type(name)))
        else:
            msg = "name %s (type %s) could not be converted to a string"
            raise AttributeError(msg % (name, type(name)))

    if name[0].isdigit():
        return False, "Variables cannot begin with numeric character"
    if _nonalphanumeric_re.search(name):
        return False, "Variables cannot contain punctuation or spaces"
    return True, ''


def _getExcelCellName(col, row):
    """Returns the excel cell name for a row and column (zero-indexed)

    >>> _getExcelCellName(0,0)
    'A1'
    >>> _getExcelCellName(2,1)
    'C2'
    """
    # BEWARE - openpyxl uses indexing at 1, to fit with Excel
    return "%s%i" % (get_column_letter(col + 1), row + 1)
